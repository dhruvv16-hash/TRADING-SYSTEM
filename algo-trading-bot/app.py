import os
import math
import time
import logging
from flask import Flask, request, jsonify, render_template, make_response
from config import Config
from delta_client import DeltaClient
from models import db, Account, GlobalSetting, TradeLog, Strategy, StrategyState, LocalSignalLog, AIInsight
import requests

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# Validate configurations on start (unless running tests or syntax checks)
if os.getenv("FLASK_ENV") != "testing":
    try:
        Config.validate()
        logger.info("Configuration validated successfully.")
    except ValueError as e:
        logger.error(f"Configuration error: {e}")
        logger.warning("Bot is starting, but missing environment variables! Please configure before trading.")

# Initialize database
# Use DATABASE_URL from Render or local .env. The tables are prefixed with bot_ 
# so it's safe to share the database with the Next.js app.
database_url = os.getenv("DATABASE_URL")
if database_url:
    # Render's database URL might start with postgres:// which SQLAlchemy 2.0 deprecated, so fix it
    if database_url.startswith("postgres://"):
        database_url = database_url.replace("postgres://", "postgresql://", 1)
    app.config["SQLALCHEMY_DATABASE_URI"] = database_url
else:
    app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///local.db"

app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_pre_ping": True,
    "pool_recycle": 280
}

db.init_app(app)

_db_pool_initialized = False

@app.before_request
def initialize_db_connections():
    global _db_pool_initialized
    if not _db_pool_initialized:
        try:
            db_uri = app.config.get("SQLALCHEMY_DATABASE_URI") or ""
            if "sqlite" not in db_uri:
                db.engine.dispose()
        except Exception as e:
            logger.warning(f"Error disposing database engine pool on first request: {e}")
            
        # Spawn daemon threads after disposing the engine pool on first request
        if os.getenv("FLASK_ENV") != "testing":
            init_email_listener()
            init_strategy_runner()
            
        _db_pool_initialized = True

@app.before_request
def track_host_url():
    app.config["LAST_HOST_URL"] = request.host_url

@app.after_request
def add_security_headers(response):
    response.headers["Content-Security-Policy"] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' 'unsafe-eval' https://cdn.jsdelivr.net; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "img-src 'self' data:; "
        "connect-src 'self';"
    )
    response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains; preload"
    # response.headers["X-Frame-Options"] = "SAMEORIGIN"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "geolocation=(), microphone=(), camera=()"
    return response

@app.route("/robots.txt")
def robots_txt():
    host = request.host_url
    content = f"User-agent: *\nAllow: /\nSitemap: {host}sitemap.xml"
    response = make_response(content)
    response.headers["Content-Type"] = "text/plain"
    return response

@app.route("/sitemap.xml")
def sitemap_xml():
    host = request.host_url
    content = f"""<?xml version="1.0" encoding="UTF-8"?>
<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">
    <url>
        <loc>{host}</loc>
        <lastmod>2026-06-20</lastmod>
        <changefreq>daily</changefreq>
        <priority>1.0</priority>
    </url>
</urlset>"""
    response = make_response(content)
    response.headers["Content-Type"] = "application/xml"
    return response

# Initialize static Delta REST Client for public symbols lookup
public_delta_client = DeltaClient(
    api_key="public",
    api_secret="public",
    base_url=Config.BASE_URL
)

def escape_telegram_html(html_str):
    import re
    # Match <a href='...'> or <a href="...">
    a_pattern = re.compile(r'(<a\s+href=[\'"]([^\'"]+)[\'"]>)', re.IGNORECASE)
    a_tags = []
    
    def a_replace(match):
        url = match.group(2)
        escaped_url = url.replace("&", "&amp;")
        rebuilt_tag = f'<a href="{escaped_url}">'
        placeholder = f"__A_TAG_PLACEHOLDER_{len(a_tags)}__"
        a_tags.append(rebuilt_tag)
        return placeholder
        
    temp_str = a_pattern.sub(a_replace, html_str)
    
    tag_replacements = {
        "<b>": "__B_OPEN__",
        "</b>": "__B_CLOSE__",
        "<i>": "__I_OPEN__",
        "</i>": "__I_CLOSE__",
        "<code>": "__CODE_OPEN__",
        "</code>": "__CODE_CLOSE__",
        "<pre>": "__PRE_OPEN__",
        "</pre>": "__PRE_CLOSE__",
        "</a>": "__A_CLOSE__",
        "<strong>": "__STRONG_OPEN__",
        "</strong>": "__STRONG_CLOSE__",
        "<em>": "__EM_OPEN__",
        "</em>": "__EM_CLOSE__",
        "<u>": "__U_OPEN__",
        "</u>": "__U_CLOSE__",
        "<ins>": "__INS_OPEN__",
        "</ins>": "__INS_CLOSE__",
        "<s>": "__S_OPEN__",
        "</s>": "__S_CLOSE__",
        "<strike>": "__STRIKE_OPEN__",
        "</strike>": "__STRIKE_CLOSE__",
        "<del>": "__DEL_OPEN__",
        "</del>": "__DEL_CLOSE__",
    }
    
    for tag, placeholder in tag_replacements.items():
        temp_str = re.sub(re.escape(tag), placeholder, temp_str, flags=re.IGNORECASE)
        
    # Escape raw text characters
    temp_str = temp_str.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    
    for tag, placeholder in tag_replacements.items():
        temp_str = temp_str.replace(placeholder, tag)
        
    for idx, original_tag in enumerate(a_tags):
        placeholder = f"__A_TAG_PLACEHOLDER_{idx}__"
        temp_str = temp_str.replace(placeholder, original_tag)
        
    return temp_str

def send_notification(title, message, status_color=3447003, symbol=None):
    import requests
    # Ensure this doesn't run during testing to avoid making external requests
    if os.getenv("FLASK_ENV") == "testing":
        return {"telegram": {"success": True}, "discord": {"success": True}}
    results = {"telegram": {"success": True}, "discord": {"success": True}}
    try:
        keys = ["telegram_enabled", "telegram_token", "telegram_chat_id", "discord_enabled", "discord_webhook_url", "passphrase"]
        settings = GlobalSetting.query.filter(GlobalSetting.key.in_(keys)).all()
        s_dict = {s.key: s.value for s in settings}
        
        telegram_enabled = s_dict.get("telegram_enabled")
        telegram_token = s_dict.get("telegram_token")
        telegram_chat_id = s_dict.get("telegram_chat_id")
        
        discord_enabled = s_dict.get("discord_enabled")
        discord_webhook_url = s_dict.get("discord_webhook_url")
        
        passphrase = s_dict.get("passphrase") or Config.PASSPHRASE
        import hashlib
        token = hashlib.sha256(passphrase.encode('utf-8')).hexdigest()[:16]
        
        host_url = app.config.get("LAST_HOST_URL", "http://localhost:5000/")
        if not host_url.endswith("/"):
            host_url += "/"
            
        quick_actions_html = ""
        is_trade_alert = "Trade Alert" in title or "Strategy" in title or "Circuit Breaker" in title or symbol is not None
        if is_trade_alert:
            halt_url = f"{host_url}api/actions/trigger?action=halt&token={token}"
            quick_actions_html = f"\n\n⚡ <b>Quick Actions:</b>\n• <a href='{halt_url}'>Halt All Trading</a>"
            if symbol:
                close_url = f"{host_url}api/actions/trigger?action=close&symbol={symbol}&token={token}"
                quick_actions_html += f"\n• <a href='{close_url}'>Close Position ({symbol})</a>"
                
        full_message = f"{message}{quick_actions_html}"

        # Discord Embed
        if discord_enabled and discord_enabled.lower() == "true" and discord_webhook_url:
            import re
            discord_msg = full_message
            discord_msg = re.sub(r'<a\s+href=[\'"]([^\'"]+)[\'"]>([^<]+)</a>', r'[\2](\1)', discord_msg)
            discord_msg = discord_msg.replace("<b>", "**").replace("</b>", "**").replace("<pre>", "```").replace("</pre>", "```")
            payload = {
                "embeds": [{
                    "title": title,
                    "description": discord_msg,
                    "color": int(status_color),
                    "timestamp": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
                }]
            }
            try:
                res = requests.post(discord_webhook_url, json=payload, timeout=5)
                if res.status_code >= 400:
                    logger.error(f"Discord notification HTTP error: {res.status_code} - {res.text}")
                    results["discord"] = {"success": False, "error": f"HTTP {res.status_code}: {res.text}"}
                else:
                    results["discord"] = {"success": True}
            except Exception as e:
                logger.error(f"Failed to post to Discord webhook: {e}")
                results["discord"] = {"success": False, "error": str(e)}

        # Telegram Message
        if telegram_enabled and telegram_enabled.lower() == "true" and telegram_token and telegram_chat_id:
            url = f"https://api.telegram.org/bot{telegram_token}/sendMessage"
            telegram_text = f"<b>{title}</b>\n\n{full_message}"
            try:
                telegram_text = escape_telegram_html(telegram_text)
            except Exception as esc_err:
                logger.error(f"Failed to escape Telegram HTML: {esc_err}")
            
            payload = {
                "chat_id": telegram_chat_id,
                "text": telegram_text,
                "parse_mode": "HTML"
            }
            try:
                res = requests.post(url, json=payload, timeout=5)
                if res.status_code >= 400:
                    logger.error(f"Telegram notification HTTP error: {res.status_code} - {res.text}")
                    results["telegram"] = {"success": False, "error": f"HTTP {res.status_code}: {res.text}"}
                else:
                    results["telegram"] = {"success": True}
            except Exception as e:
                logger.error(f"Failed to send to Telegram bot: {e}")
                results["telegram"] = {"success": False, "error": str(e)}
    except Exception as e:
        logger.error(f"Error executing send_notification: {e}")
        results["error"] = str(e)
    return results

# Track the global online status of Delta Exchange API
last_api_status = {"online": True, "last_alert_time": 0}

def send_email_alert(subject, html_body):
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    # Ensure this doesn't run during testing to avoid making external requests
    if os.getenv("FLASK_ENV") == "testing":
        return True

    try:
        keys = ["email_enabled", "email_address", "email_password", "imap_host"]
        s_dict = {}
        try:
            settings = GlobalSetting.query.filter(GlobalSetting.key.in_(keys)).all()
            s_dict = {s.key: s.value for s in settings}
        except Exception as e:
            logger.debug(f"Failed to load email settings from DB: {e}")

        enabled = s_dict.get("email_enabled") or os.getenv("EMAIL_ENABLED")
        if not enabled or enabled.lower() != "true":
            logger.info("Email alerts not enabled globally.")
            return False

        email_address_s_val = s_dict.get("email_address") or os.getenv("EMAIL_ADDRESS")
        email_password_s_val = s_dict.get("email_password") or os.getenv("EMAIL_PASSWORD")
        imap_host_s_val = s_dict.get("imap_host") or os.getenv("IMAP_HOST")

        if not email_address_s_val or not email_password_s_val:
            logger.warning("Email address or password not configured. Cannot send email alert.")
            return False

        email_address = email_address_s_val
        email_password = email_password_s_val
        imap_host = imap_host_s_val if imap_host_s_val else "imap.gmail.com"

        # Determine SMTP Host
        smtp_host = "smtp.gmail.com"
        smtp_port = 587
        
        host_lower = imap_host.lower()
        if "gmail" in host_lower:
            smtp_host = "smtp.gmail.com"
            smtp_port = 587
        elif "outlook" in host_lower or "office365" in host_lower or "live.com" in host_lower:
            smtp_host = "smtp.office365.com"
            smtp_port = 587
        elif "yahoo" in host_lower:
            smtp_host = "smtp.mail.yahoo.com"
            smtp_port = 465
        else:
            smtp_host = imap_host.replace("imap", "smtp")
            smtp_port = 587

        msg = MIMEMultipart()
        msg["From"] = email_address
        msg["To"] = email_address
        msg["Subject"] = subject
        msg.attach(MIMEText(html_body, "html"))

        if smtp_port == 465:
            server = smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=10)
        else:
            server = smtplib.SMTP(smtp_host, smtp_port, timeout=10)
            server.starttls()

        server.login(email_address, email_password)
        server.sendmail(email_address, email_address, msg.as_string())
        server.quit()
        logger.info(f"Successfully sent email alert: {subject}")
        return True
    except Exception as e:
        logger.error(f"Failed to send email alert: {e}")
        return False

def verify_api_connectivity():
    """
    Pings Delta Exchange public endpoint to verify connectivity.
    If the API transitions to offline state, sends an email alert to the user.
    """
    global last_api_status
    import requests
    import time

    if os.getenv("FLASK_ENV") == "testing":
        return True

    base_url = Config.BASE_URL
    url = f"{base_url}/v2/products"

    online = False
    error_msg = ""
    try:
        res = requests.get(url, timeout=5)
        if res.status_code == 200:
            online = True
        else:
            error_msg = f"HTTP status code {res.status_code} - {res.text}"
    except Exception as e:
        error_msg = str(e)

    now = time.time()
    if not online:
        logger.warning(f"Delta Exchange API connectivity check failed: {error_msg}")
        if last_api_status["online"] or (now - last_api_status["last_alert_time"] > 3600):
            last_api_status["online"] = False
            last_api_status["last_alert_time"] = now
            
            subject = "⚠️ Alert: Delta Exchange API is OFFLINE / Unresponsive"
            body = f"""
            <h3>Delta Exchange API Connectivity Alert</h3>
            <p>The Trading Bot detected that the Delta Exchange API at <b>{base_url}</b> is offline or unresponsive.</p>
            <p><b>Error Details:</b></p>
            <pre style="background: #f4f4f4; padding: 1rem; border-radius: 4px;">{error_msg}</pre>
            <p>The bot will automatically retry connecting in the background. Webhook signals and local strategies may fail to execute while the exchange is offline.</p>
            <hr>
            <p>Sent by Delta Bot Admin</p>
            """
            send_email_alert(subject, body)
            send_notification(subject, f"The Trading Bot detected that the Delta Exchange API at <b>{base_url}</b> is offline or unresponsive.\n\nError details:\n<pre>{error_msg}</pre>", status_color=15680580)
    else:
        if not last_api_status["online"]:
            logger.info("Delta Exchange API connection restored.")
            last_api_status["online"] = True
            
            subject = "🟢 Resolved: Delta Exchange API is back ONLINE"
            body = f"""
            <h3>Delta Exchange API Restored</h3>
            <p>The connection to Delta Exchange API at <b>{base_url}</b> has been successfully restored. All systems are operating normally.</p>
            <hr>
            <p>Sent by Delta Bot Admin</p>
            """
            send_email_alert(subject, body)
            send_notification(subject, f"The connection to Delta Exchange API at <b>{base_url}</b> has been successfully restored. All systems are operating normally.", status_color=1096065)

    return online

# ----------------- EMAIL DOUBLE-VERIFICATION INTEGRATION -----------------

def parse_email_signal(body, subject):
    import re
    import json
    
    # Heuristics:
    # 1. Look for JSON in the body
    json_match = re.search(r'(\{.*?\})', body, re.DOTALL)
    if json_match:
        try:
            data = json.loads(json_match.group(1))
            ticker = data.get("ticker") or data.get("symbol")
            action = data.get("action", "").lower()
            quantity = data.get("quantity") or data.get("qty")
            if ticker and action:
                return ticker, action, quantity
        except Exception:
            pass
            
    # 2. Key-Value pairs
    ticker = None
    action = None
    quantity = None
    
    ticker_match = re.search(r'(?:ticker|symbol)\s*[:=]\s*["\']?([a-zA-Z0-9_:\.\/]+)["\']?', body, re.IGNORECASE)
    if ticker_match:
        ticker = ticker_match.group(1)
        
    action_match = re.search(r'(?:action|side)\s*[:=]\s*["\']?(buy|sell|close_long|close_short)["\']?', body, re.IGNORECASE)
    if action_match:
        action = action_match.group(1).lower()
        
    qty_match = re.search(r'(?:quantity|qty)\s*[:=]\s*["\']?([0-9\.]+)["\']?', body, re.IGNORECASE)
    if qty_match:
        try:
            quantity = float(qty_match.group(1))
        except ValueError:
            pass
            
    # 3. Subject line parsing fallback
    if not ticker or not action:
        words = re.findall(r'\b[a-zA-Z0-9_\:\.\/]+\b', subject + " " + body)
        for word in words:
            word_lower = word.lower()
            if word_lower in ["buy", "sell", "close_long", "close_short"]:
                action = word_lower
            elif ("usd" in word_lower or "btc" in word_lower or "eth" in word_lower) and len(word) >= 5:
                if not ticker:
                    ticker = word
                    
    return ticker, action, quantity

def check_position_matches_action(account_data, ticker, action):
    try:
        product = public_delta_client.get_product_by_symbol(ticker)
        if not product:
            return False
            
        product_id = product.get("id")
        
        client = DeltaClient(
            api_key=account_data["api_key"],
            api_secret=account_data["api_secret"],
            base_url=Config.BASE_URL
        )
        
        pos = client.get_position(product_id)
        size = 0.0
        side = ""
        if pos:
            try:
                size = float(pos.get("size", 0.0))
                side = pos.get("side", "").lower()
            except (ValueError, TypeError):
                pass
                
        if action == "buy":
            return size > 0 or side == "buy"
        elif action == "sell":
            return size < 0 or side == "sell"
        elif action == "close_long":
            return size <= 0
        elif action == "close_short":
            return size >= 0
            
    except Exception as e:
        logger.error(f"Error checking position on Delta: {e}")
        return False
        
    return False

def email_polling_loop():
    import imaplib
    import email
    import time
    import socket
    
    logger.info("Email polling background worker started.")
    while True:
        try:
            # Poll every 60 seconds
            time.sleep(60)
            
            enabled = False
            imap_host = ""
            imap_port = 993
            email_address = ""
            email_password = ""
            email_sender = "noreply@tradingview.com"
            email_subject = "TradingView Alert"
            accounts_data = []
            
            # 1. Fetch settings inside a short-lived DB context and release connection immediately
            with app.app_context():
                keys = ["email_enabled", "imap_host", "imap_port", "email_address", "email_password", "email_sender", "email_subject"]
                settings = GlobalSetting.query.filter(GlobalSetting.key.in_(keys)).all()
                s_dict = {s.key: s.value for s in settings}
                
                enabled_setting = s_dict.get("email_enabled")
                if enabled_setting == "true":
                    enabled = True
                    imap_host = s_dict.get("imap_host", "")
                    
                    imap_port_val = s_dict.get("imap_port")
                    if imap_port_val:
                        try:
                            imap_port = int(imap_port_val)
                        except ValueError:
                            imap_port = 993
                    else:
                        imap_port = 993
                        
                    email_address = s_dict.get("email_address", "")
                    email_password = s_dict.get("email_password", "")
                    email_sender = s_dict.get("email_sender", "noreply@tradingview.com")
                    email_subject = s_dict.get("email_subject", "TradingView Alert")
                        
                    active_accounts = Account.query.filter_by(is_active=True).all()
                    for account in active_accounts:
                        accounts_data.append({
                            "id": account.id,
                            "name": account.name,
                            "api_key": account.api_key,
                            "api_secret": account.api_secret,
                            "leverage": account.leverage,
                            "balance_buffer_pct": account.balance_buffer_pct,
                            "sizing_type": account.sizing_type,
                            "fixed_amount": account.fixed_amount
                        })
                        
                    if not accounts_data and Config.API_KEY and Config.API_SECRET:
                        accounts_data = [{
                            "id": 0,
                            "name": "Environment Default",
                            "api_key": Config.API_KEY,
                            "api_secret": Config.API_SECRET,
                            "leverage": Config.DEFAULT_LEVERAGE,
                            "balance_buffer_pct": Config.BALANCE_BUFFER_PCT * 100.0
                        }]
            
            if not enabled or not imap_host or not email_address or not email_password:
                continue
                
            # 2. Perform IMAP operations outside the database context
            logger.info(f"Connecting to IMAP {imap_host}:{imap_port} for {email_address}...")
            
            # Set a global socket timeout of 15 seconds to prevent hanging indefinitely
            original_timeout = socket.getdefaulttimeout()
            socket.setdefaulttimeout(15)
            try:
                mail = imaplib.IMAP4_SSL(imap_host, imap_port)
                mail.login(email_address, email_password)
                mail.select("inbox")
                
                status, messages = mail.search(None, 'UNSEEN')
                if status != "OK":
                    mail.logout()
                    continue
                    
                mail_ids = messages[0].split()
                if not mail_ids:
                    mail.logout()
                    continue
                    
                logger.info(f"Detected {len(mail_ids)} unread emails. Reconciling signals...")
                
                for mail_id in mail_ids:
                    res, msg_data = mail.fetch(mail_id, '(RFC822)')
                    if res != "OK":
                        continue
                        
                    for response_part in msg_data:
                        if isinstance(response_part, tuple):
                            raw_email = response_part[1]
                            msg = email.message_from_bytes(raw_email)
                            
                            sender_header = msg.get("From", "")
                            subject_header = msg.get("Subject", "")
                            
                            if email_sender.lower() not in sender_header.lower():
                                continue
                            if email_subject.lower() not in subject_header.lower():
                                continue
                                
                            logger.info(f"Reconciling email signal: '{subject_header}' from '{sender_header}'")
                            
                            body = ""
                            if msg.is_multipart():
                                for part in msg.walk():
                                    content_type = part.get_content_type()
                                    content_disposition = str(part.get("Content-Disposition"))
                                    if content_type == "text/plain" and "attachment" not in content_disposition:
                                        payload = part.get_payload(decode=True)
                                        if payload:
                                            body += payload.decode('utf-8', errors='ignore')
                            else:
                                payload = msg.get_payload(decode=True)
                                if payload:
                                    body += payload.decode('utf-8', errors='ignore')
                                    
                            ticker, action, quantity = parse_email_signal(body, subject_header)
                            if not ticker or not action:
                                logger.warning("Failed to extract ticker/action from email body.")
                                continue
                                
                            mail.store(mail_id, '+FLAGS', '\\Seen')
                            logger.info(f"Email marked as read. Extracted signal: {action} {ticker} (Quantity: {quantity})")
                            
                            if not accounts_data:
                                logger.warning("No accounts available to check position for email signal.")
                                continue
                                
                            # Re-enter DB context only to perform database writes
                            with app.app_context():
                                has_match = check_position_matches_action(accounts_data[0], ticker, action)
                                if has_match:
                                    logger.info(f"Double-Verification: Matching position for {ticker} ({action}) already exists. Skipping.")
                                    log_entry = TradeLog(
                                        ticker=ticker,
                                        action=action,
                                        source="email_fallback",
                                        status="verified",
                                        details="Verified: Position already matches the signal on Delta Exchange."
                                    )
                                    db.session.add(log_entry)
                                    db.session.commit()
                                else:
                                    logger.warning(f"Double-Verification FAILED: No active position matches {action} {ticker} on Delta. Executing fallback...")
                                    payload_data = {"quantity": quantity} if quantity is not None else None
                                    execute_trades_background(accounts_data, ticker, action, source="email_fallback", payload=payload_data)
                mail.logout()
            finally:
                socket.setdefaulttimeout(original_timeout)
        except Exception as e:
            logger.exception(f"Error in email polling iteration: {e}")

_email_thread_started = False

def init_email_listener():
    global _email_thread_started
    if _email_thread_started:
        return
        
    import threading
    thread = threading.Thread(target=email_polling_loop, daemon=True)
    thread.start()
    _email_thread_started = True
    logger.info("Spawned daemon thread for email polling.")

_strategy_thread_started = False

def init_strategy_runner():
    global _strategy_thread_started
    if _strategy_thread_started:
        return
        
    import threading
    from strategy_runner import strategy_runner_loop
    thread = threading.Thread(target=strategy_runner_loop, args=(app,), daemon=True)
    thread.start()
    _strategy_thread_started = True
    logger.info("Spawned daemon thread for local strategy runner.")

@app.route("/health", methods=["GET"])
def health():
    """Health check endpoint to keep the bot awake (e.g. via UptimeRobot)."""
    return jsonify({"status": "healthy", "service": "delta-webhook-bot"}), 200

@app.route("/", methods=["GET"])
@app.route("/dashboard", methods=["GET"])
def dashboard():
    """Serves the bot configuration panel dashboard."""
    return render_template("dashboard.html")

@app.route("/api/ping", methods=["GET"])
def api_ping():
    return jsonify({"status": "pong"})

# ----------------- ADMIN SETTINGS ENDPOINTS -----------------

@app.route("/api/settings", methods=["GET"])
def get_settings():
    keys = [
        "passphrase", "telegram_enabled", "telegram_token", "telegram_chat_id", 
        "discord_enabled", "discord_webhook_url", "local_bot_dry_run",
        "local_bot_symbols", "local_strategy_ce_length", "local_strategy_ce_mult",
        "local_strategy_zlsma_length", "local_strategy_vol_length", "local_strategy_vol_mult",
        "local_strategy_resolution", "local_strategy_tp1_rr", "local_strategy_tp2_rr",
        "local_strategy_use_be", "local_strategy_use_liq_exit", "local_strategy_use_zlsma_exit",
        "local_strategy_type", "local_strategy_zl_length", "local_strategy_zl_mult",
        "local_strategy_ai_speed", "local_strategy_ai_atr_len", "local_strategy_ai_atr_mult",
        "ai_api_key", "ai_provider", "ai_model"
    ]
    settings = GlobalSetting.query.filter(GlobalSetting.key.in_(keys)).all()
    s_dict = {s.key: s.value for s in settings}
    
    passphrase_val = s_dict.get("passphrase", Config.PASSPHRASE)
    telegram_enabled_val = s_dict.get("telegram_enabled", "false")
    telegram_token_val = s_dict.get("telegram_token", "")
    telegram_chat_id_val = s_dict.get("telegram_chat_id", "")
    discord_enabled_val = s_dict.get("discord_enabled", "false")
    discord_webhook_url_val = s_dict.get("discord_webhook_url", "")
    local_bot_dry_run_val = s_dict.get("local_bot_dry_run", "true")
    
    base_url = Config.BASE_URL.lower()
    if "testnet" in base_url:
        ws_url = "wss://api.testnet.delta.exchange/v2/websocket"
    elif "india" in base_url:
        ws_url = "wss://api.india.delta.exchange/v2/websocket"
    else:
        ws_url = "wss://api.delta.exchange/v2/websocket"
        
    return jsonify({
        "passphrase": passphrase_val,
        "telegram_enabled": telegram_enabled_val,
        "telegram_token": telegram_token_val,
        "telegram_chat_id": telegram_chat_id_val,
        "discord_enabled": discord_enabled_val,
        "discord_webhook_url": discord_webhook_url_val,
        "local_bot_dry_run": (local_bot_dry_run_val == "true"),
        "local_bot_symbols": s_dict.get("local_bot_symbols", "ETHUSD.P"),
        "local_strategy_type": s_dict.get("local_strategy_type", "chandelier_exit"),
        "local_strategy_ce_length": int(s_dict.get("local_strategy_ce_length", 22)),
        "local_strategy_ce_mult": float(s_dict.get("local_strategy_ce_mult", 3.0)),
        "local_strategy_zlsma_length": int(s_dict.get("local_strategy_zlsma_length", 32)),
        "local_strategy_vol_length": int(s_dict.get("local_strategy_vol_length", 20)),
        "local_strategy_vol_mult": float(s_dict.get("local_strategy_vol_mult", 1.15)),
        "local_strategy_zl_length": int(s_dict.get("local_strategy_zl_length", 70)),
        "local_strategy_zl_mult": float(s_dict.get("local_strategy_zl_mult", 1.2)),
        "local_strategy_ai_speed": int(s_dict.get("local_strategy_ai_speed", 14)),
        "local_strategy_ai_atr_len": int(s_dict.get("local_strategy_ai_atr_len", 14)),
        "local_strategy_ai_atr_mult": float(s_dict.get("local_strategy_ai_atr_mult", 2.0)),
        "local_strategy_resolution": s_dict.get("local_strategy_resolution", "5"),
        "local_strategy_tp1_rr": float(s_dict.get("local_strategy_tp1_rr", 1.5)),
        "local_strategy_tp2_rr": float(s_dict.get("local_strategy_tp2_rr", 3.0)),
        "local_strategy_use_be": s_dict.get("local_strategy_use_be", "true") == "true",
        "local_strategy_use_liq_exit": s_dict.get("local_strategy_use_liq_exit", "true") == "true",
        "local_strategy_use_zlsma_exit": s_dict.get("local_strategy_use_zlsma_exit", "true") == "true",
        "ws_url": ws_url,
        "ai_api_key": s_dict.get("ai_api_key", ""),
        "ai_provider": s_dict.get("ai_provider", "gemini"),
        "ai_model": s_dict.get("ai_model", "gemini-1.5-flash")
    })
 
@app.route("/api/settings", methods=["POST"])
def save_settings():
    data = request.get_json(silent=True) or {}
    
    keys = [
        "passphrase", "telegram_enabled", "telegram_token", "telegram_chat_id", 
        "discord_enabled", "discord_webhook_url", "local_bot_dry_run",
        "local_bot_symbols", "local_strategy_ce_length", "local_strategy_ce_mult",
        "local_strategy_zlsma_length", "local_strategy_vol_length", "local_strategy_vol_mult",
        "local_strategy_resolution", "local_strategy_tp1_rr", "local_strategy_tp2_rr",
        "local_strategy_use_be", "local_strategy_use_liq_exit", "local_strategy_use_zlsma_exit",
        "local_strategy_type", "local_strategy_zl_length", "local_strategy_zl_mult",
        "local_strategy_ai_speed", "local_strategy_ai_atr_len", "local_strategy_ai_atr_mult",
        "ai_api_key", "ai_provider", "ai_model"
    ]
    for key in keys:
        if key in data:
            if key == "local_bot_dry_run":
                val = str(data[key]).lower()
            elif key in ["local_strategy_use_be", "local_strategy_use_liq_exit", "local_strategy_use_zlsma_exit"]:
                val = "true" if data[key] is True or str(data[key]).lower() == "true" else "false"
            else:
                val = str(data[key])
                
            setting = GlobalSetting.query.filter_by(key=key).first()
            if setting:
                setting.value = val
            else:
                setting = GlobalSetting(key=key, value=val)
                db.session.add(setting)
    
    db.session.commit()
    return jsonify({"status": "success", "message": "Settings saved"})

@app.route("/api/actions/trigger", methods=["GET"])
def trigger_action():
    action = request.args.get("action")
    symbol = request.args.get("symbol")
    token = request.args.get("token")
    
    # 1. Verify token
    passphrase_setting = GlobalSetting.query.filter_by(key="passphrase").first()
    passphrase = passphrase_setting.value if passphrase_setting else Config.PASSPHRASE
    import hashlib
    expected_token = hashlib.sha256(passphrase.encode('utf-8')).hexdigest()[:16]
    
    if not token or token != expected_token:
        return "Unauthorized: Invalid or missing action token", 401
        
    if action == "halt":
        # Halt all trading by setting is_circuit_broken to True on all accounts
        try:
            active_accounts = Account.query.filter_by(is_active=True).all()
            for acc in active_accounts:
                acc.is_circuit_broken = True
            db.session.commit()
            
            logger.warning("Halt action triggered remotely. Circuit breakers tripped for all active accounts.")
            send_notification("🚨 Remote Action Triggered: Halt Trading", "Trading has been halted remotely across all accounts. Circuit breakers are tripped.", 15549011)
            return "Halted: All circuit breakers tripped successfully."
        except Exception as e:
            logger.exception(f"Error in remote halt action: {e}")
            return f"Error: {e}", 500
            
    elif action == "close":
        # Close positions
        try:
            active_accounts = Account.query.filter_by(is_active=True).all()
            details = []
            for acc in active_accounts:
                client = DeltaClient(
                    api_key=acc.api_key,
                    api_secret=acc.api_secret,
                    base_url=Config.BASE_URL
                )
                if symbol:
                    # Close specific symbol position
                    product = client.get_product_by_symbol(symbol)
                    if product:
                        p_id = product.get("id")
                        pos = client.get_position(p_id)
                        if pos:
                            pos_size = abs(int(float(pos.get("size", 0))))
                            if pos_size > 0:
                                entry_px = float(pos.get("entry_price") or pos.get("avg_entry_price") or 0.0)
                                pos_side = pos.get("side", "").lower()
                                is_long = (pos_side == "buy" or float(pos.get("size", 0)) > 0)
                                close_side = "sell" if is_long else "buy"
                                res = client.place_order(
                                    product_id=p_id,
                                    size=pos_size,
                                    side=close_side,
                                    order_type="market_order",
                                    reduce_only=True
                                )
                                pnl_val = 0.0
                                if res.get("success"):
                                    order_res = res.get("result", {})
                                    try:
                                        exit_px = float(order_res.get("average_fill_price") or 0.0)
                                    except (ValueError, TypeError):
                                        exit_px = 0.0
                                    
                                    if exit_px <= 0:
                                        ticker_data = client.get_ticker(symbol)
                                        exit_px = float(ticker_data.get("mark_price") or ticker_data.get("last_price") or 0.0)
                                        
                                    if entry_px > 0 and exit_px > 0:
                                        direction = 1 if is_long else -1
                                        contract_value = 0.01
                                        try:
                                            contract_value = float(product.get("contract_value", "0.01"))
                                        except Exception:
                                            pass
                                        pnl_val = (exit_px - entry_px) * pos_size * contract_value * direction
                                    
                                    pnl_str = f"PnL: {pnl_val:+.2f} USD"
                                    details.append(f"{acc.name}: Closed {pos_size} lots of {symbol} ({pnl_str})")
                                    
                                    # Reset virtual position state if any
                                    states = StrategyState.query.filter_by(account_id=acc.id, symbol=symbol).all()
                                    for state_db in states:
                                        state_db.position_size = 0.0
                                        state_db.entry_price = None
                                        state_db.current_sl = None
                                        state_db.tp1_price = None
                                        state_db.tp2_price = None
                                        state_db.tp1_hit = False
                                        state_db.tp2_hit = False
                                else:
                                    details.append(f"{acc.name}: Failed to close {symbol}: {res.get('message', 'API Error')}")
                            else:
                                details.append(f"{acc.name}: No open position for {symbol}")
                        else:
                            details.append(f"{acc.name}: No open position for {symbol}")
                else:
                    # Close all positions
                    closed_list = close_all_positions(client)
                    if closed_list:
                        details.append(f"{acc.name}: {', '.join(closed_list)}")
                        # Reset all StrategyStates
                        states = StrategyState.query.filter_by(account_id=acc.id).all()
                        for state_db in states:
                            state_db.position_size = 0.0
                            state_db.entry_price = None
                            state_db.current_sl = None
                            state_db.tp1_price = None
                            state_db.tp2_price = None
                            state_db.tp1_hit = False
                            state_db.tp2_hit = False
                    else:
                        details.append(f"{acc.name}: No open positions found")
            
            db.session.commit()
            details_str = "\n".join(details)
            send_notification("⚡ Remote Action Triggered: Close Position", f"Positions close triggered remotely:\n\n{details_str}", 1096065)
            return f"Closed: {details_str}"
        except Exception as e:
            logger.exception(f"Error in remote close action: {e}")
            return f"Error: {e}", 500
            
    else:
        return f"Unknown action: {action}", 400

@app.route("/api/notifications/test", methods=["POST"])
def test_notification():
    title = "🔔 Delta Bot Alert: Test Connection"
    message = "Your Telegram and Discord alert integration was configured and tested successfully!"
    results = send_notification(title, message, 3447003)
    
    # Check if enabled integrations failed and report detailed error messages
    telegram_enabled = GlobalSetting.query.filter_by(key="telegram_enabled").first()
    discord_enabled = GlobalSetting.query.filter_by(key="discord_enabled").first()
    
    errors = []
    if telegram_enabled and telegram_enabled.value.lower() == "true":
        tel_res = results.get("telegram", {})
        if not tel_res.get("success"):
            errors.append(f"Telegram: {tel_res.get('error', 'unknown error')}")
            
    if discord_enabled and discord_enabled.value.lower() == "true":
        disc_res = results.get("discord", {})
        if not disc_res.get("success"):
            errors.append(f"Discord: {disc_res.get('error', 'unknown error')}")
            
    if errors:
        return jsonify({"status": "error", "message": " | ".join(errors)}), 400
        
    return jsonify({"status": "success", "message": "Test notification dispatched"})

@app.route("/api/simulate-webhook", methods=["POST"])
def simulate_webhook():
    try:
        data = request.get_json(silent=True) or {}
        ticker = data.get("ticker")
        action = data.get("action")
        passphrase = data.get("passphrase")
        live_execute = data.get("live_execute", False)
        
        # 1. Validation
        if not ticker or not action or not passphrase:
            return jsonify({"status": "error", "message": "Missing required fields: ticker, action, passphrase"}), 400
            
        passphrase_setting = GlobalSetting.query.filter_by(key="passphrase").first()
        configured_passphrase = passphrase_setting.value if passphrase_setting else Config.PASSPHRASE
        if passphrase != configured_passphrase:
            return jsonify({"status": "error", "message": "Invalid passphrase"}), 401
            
        action = action.lower()
        if action not in ["buy", "sell", "close_long", "close_short"]:
            return jsonify({"status": "error", "message": f"Invalid action: {action}"}), 400

        # Fetch active accounts
        active_accounts = Account.query.filter_by(is_active=True).all()
        if not active_accounts:
            if Config.API_KEY and Config.API_SECRET:
                fallback_account = Account(
                    id=0,
                    name="Environment Default",
                    api_key=Config.API_KEY,
                    api_secret=Config.API_SECRET,
                    leverage=Config.DEFAULT_LEVERAGE,
                    balance_buffer_pct=Config.BALANCE_BUFFER_PCT * 100.0,
                    sizing_type="percentage",
                    fixed_amount=10.0,
                    is_active=True
                )
                active_accounts = [fallback_account]
            else:
                return jsonify({"status": "error", "message": "No active trading accounts configured"}), 400

        # Resolve product
        product = public_delta_client.get_product_by_symbol(ticker)
        if not product:
            return jsonify({"status": "error", "message": f"Ticker '{ticker}' not found on Delta Exchange"}), 404

        symbol = product.get("symbol")
        product_id = product.get("id")
        contract_value = float(product.get("contract_value", "0.01"))
        
        if live_execute:
            # Trigger live background order execution
            accounts_data = [acc.to_dict() for acc in active_accounts]
            for acc_dict in accounts_data:
                # Add unmasked keys for execution
                acc_db = Account.query.get(acc_dict["id"]) if acc_dict["id"] != 0 else None
                acc_dict["api_key"] = acc_db.api_key if acc_db else Config.API_KEY
                acc_dict["api_secret"] = acc_db.api_secret if acc_db else Config.API_SECRET
            
            import threading
            threading.Thread(target=execute_trades_background, args=(accounts_data, ticker, action, "sandbox_live", data)).start()
            return jsonify({
                "status": "success",
                "message": "Live trade execution started in background",
                "details": f"Symbol: {symbol}, Action: {action.upper()}"
            })

        # Dry Run Simulation
        simulation_logs = []
        for acc in active_accounts:
            acc_name = acc.name
            sim_log = {
                "account": acc_name,
                "success": True,
                "message": ""
            }
            try:
                # Mock client for balance & tickers in test mode, otherwise use actual credentials
                client = DeltaClient(
                    api_key=acc.api_key,
                    api_secret=acc.api_secret,
                    base_url=Config.BASE_URL
                )
                
                # Fetch balance
                if os.getenv("FLASK_ENV") == "testing" or acc.api_key == "key1":
                    balance, asset = 100.0, "USD"
                    price = 2000.0
                else:
                    try:
                        balance, asset = client.get_available_balance()
                        ticker_data = client.get_ticker(symbol)
                        price = float(ticker_data.get("mark_price") or ticker_data.get("last_price") or 2000.0)
                    except Exception as client_err:
                        balance, asset = 100.0, "USD"
                        price = 2000.0
                        simulation_logs.append({
                            "account": acc_name,
                            "success": False,
                            "message": f"Could not fetch live balance/price (simulated with 100 USD @ $2000): {client_err}"
                        })
                        continue

                lot_value_usd = price * contract_value
                qty_lots = None
                sizing_desc = ""

                # Payload override
                payload_qty = data.get("quantity") or data.get("qty")
                if payload_qty is not None:
                    try:
                        qty_base = float(payload_qty)
                        qty_lots = int(math.floor(qty_base / contract_value))
                        sizing_desc = f"Payload Quantity = {qty_base} (Lots = {qty_lots})"
                    except Exception:
                        pass
                
                strategy_name = data.get("strategy") or data.get("strategy_name")
                strategy = None
                if strategy_name and acc.id != 0:
                    strategy_name = strategy_name.strip()
                    strategy = Strategy.query.filter_by(account_id=acc.id).filter(Strategy.name.ilike(strategy_name)).first()

                if qty_lots is None:
                    if strategy:
                        sizing_type = strategy.sizing_type
                        fixed_amount = strategy.fixed_amount
                        buffer_pct = strategy.balance_buffer_pct
                        leverage = strategy.leverage
                        
                        if sizing_type == "fixed":
                            if fixed_amount > balance:
                                sim_log.update({
                                    "success": False,
                                    "message": f"Simulation failed: Strategy '{strategy.name}' Fixed Margin {fixed_amount} {asset} exceeds balance {balance} {asset}."
                                })
                                simulation_logs.append(sim_log)
                                continue
                            buying_power = fixed_amount * leverage
                            sizing_desc = f"Strategy '{strategy.name}' Fixed Margin = {fixed_amount} {asset}"
                        else:
                            buying_power = balance * leverage * (buffer_pct / 100.0)
                            sizing_desc = f"Strategy '{strategy.name}' Allocation = {buffer_pct}%"
                            
                        qty_lots = int(math.floor(buying_power / lot_value_usd))
                    else:
                        if acc.sizing_type == "fixed":
                            if acc.fixed_amount > balance:
                                sim_log.update({
                                    "success": False,
                                    "message": f"Simulation failed: Fixed Margin {acc.fixed_amount} {asset} exceeds balance {balance} {asset}."
                                })
                                simulation_logs.append(sim_log)
                                continue
                            buying_power = acc.fixed_amount * acc.leverage
                            sizing_desc = f"Fixed Margin = {acc.fixed_amount} {asset}"
                        else:
                            buying_power = balance * acc.leverage * (acc.balance_buffer_pct / 100.0)
                            sizing_desc = f"Percentage Allocation = {acc.balance_buffer_pct}%"
                            
                        qty_lots = int(math.floor(buying_power / lot_value_usd))

                used_leverage = strategy.leverage if strategy else acc.leverage
                required_margin = (qty_lots * lot_value_usd) / used_leverage
                if required_margin > balance:
                    sim_log.update({
                        "success": False,
                        "message": f"Simulation failed: Margin required ({required_margin:.2f} {asset}) for {qty_lots} lots exceeds balance ({balance:.2f} {asset}) at leverage {used_leverage}x."
                    })
                elif qty_lots <= 0:
                    sim_log.update({
                        "success": False,
                        "message": "Simulation failed: Computed quantity is 0 lots. Insufficient margin."
                    })
                else:
                    sim_log.update({
                        "success": True,
                        "message": f"Simulated successfully: Would place {action.upper()} market order of size {qty_lots} lots (~{qty_lots * contract_value:.4f} base asset) on {symbol} @ mark price ${price:.2f}. Sizing rule: {sizing_desc}. Leverage: {used_leverage}x. Required margin: ${required_margin:.2f} {asset}."
                    })
            except Exception as e:
                sim_log.update({
                    "success": False,
                    "message": f"Simulation error: {str(e)}"
                })
            simulation_logs.append(sim_log)

        return jsonify({
            "status": "success",
            "simulation": True,
            "ticker": ticker,
            "action": action,
            "symbol": symbol,
            "contract_value": contract_value,
            "results": simulation_logs
        })
    except Exception as outer_err:
        logger.exception(f"Error in simulate_webhook: {outer_err}")
        return jsonify({"status": "error", "message": f"Webhook simulation failed: {str(outer_err)}"}), 500
@app.route("/api/analytics", methods=["GET"])
def get_analytics():
    try:
        # We need to compile closed trades history from the exchange
        active_accounts = Account.query.filter_by(is_active=True).all()
        if not active_accounts:
            if Config.API_KEY and Config.API_SECRET:
                fallback_account = Account(
                    id=0,
                    name="Environment Default",
                    api_key=Config.API_KEY,
                    api_secret=Config.API_SECRET,
                    leverage=Config.DEFAULT_LEVERAGE,
                    balance_buffer_pct=Config.BALANCE_BUFFER_PCT * 100.0,
                    sizing_type="percentage",
                    fixed_amount=10.0,
                    is_active=True
                )
                active_accounts = [fallback_account]
            else:
                return jsonify({
                    "status": "success", 
                    "metrics": {
                        "win_rate": 0,
                        "profit_factor": 0,
                        "sharpe_ratio": 0,
                        "recovery_factor": 0,
                        "total_trades": 0,
                        "net_profit": 0
                    }, 
                    "series_pnl": [], 
                    "series_drawdown": [], 
                    "heatmap": []
                })

        # Fetch products for symbol mapping
        products = []
        try:
            products = public_delta_client.get_products()
        except Exception as e:
            logger.warning(f"Failed to fetch products for analytics: {e}")
        product_map = {p.get("id"): p for p in products if p.get("id")}

        all_closed = []
        for account in active_accounts:
            try:
                client = DeltaClient(
                    api_key=account.api_key,
                    api_secret=account.api_secret,
                    base_url=Config.BASE_URL
                )
                closed = client.get_closed_positions(limit=150)
                for pos in closed:
                    # Calculate net PnL and timestamps
                    product_id = pos.get("product_id")
                    prod_info = product_map.get(product_id) or pos.get("product") or {}
                    symbol = prod_info.get("symbol") or f"ID:{product_id}"
                    
                    rpnl = float(pos.get("realized_pnl") or pos.get("pnl") or 0.0)
                    
                    closed_at = pos.get("closed_at")
                    closed_at_raw = 0
                    closed_at_str = ""
                    if closed_at:
                        try:
                            import datetime
                            iso_str = str(closed_at)
                            if iso_str.endswith('Z'):
                                iso_str = iso_str[:-1] + '+00:00'
                            try:
                                dt = datetime.datetime.fromisoformat(iso_str)
                                closed_at_raw = dt.timestamp()
                                closed_at_str = dt.strftime("%Y-%m-%d %H:%M:%S")
                            except ValueError:
                                t_val = float(closed_at)
                                if t_val > 1e12:
                                    t_val = t_val / 1000.0
                                if t_val > 1e11:
                                    t_val = t_val / 1000.0
                                closed_at_raw = t_val
                                dt = datetime.datetime.fromtimestamp(t_val, datetime.timezone.utc)
                                closed_at_str = dt.strftime("%Y-%m-%d %H:%M:%S")
                        except Exception:
                            closed_at_str = str(closed_at)

                    # Estimate fees
                    realized_fee_val = pos.get("fee") or pos.get("realized_fee") or pos.get("commission")
                    if realized_fee_val is not None:
                        fees = abs(float(realized_fee_val))
                    else:
                        entry_px = float(pos.get("entry_price") or pos.get("avg_entry_price") or 0)
                        exit_px = float(pos.get("close_price") or pos.get("exit_price") or 0)
                        contract_val_str = prod_info.get("contract_value") or "0.01"
                        try:
                            contract_value = float(contract_val_str)
                        except ValueError:
                            contract_value = 0.01
                        c_size = abs(float(pos.get("closed_size") or pos.get("size") or 0.0))
                        entry_notional = entry_px * c_size * contract_value
                        exit_notional = exit_px * c_size * contract_value
                        fees = (entry_notional + exit_notional) * 0.0005

                    net_pnl = rpnl - fees
                    all_closed.append({
                        "net_pnl": net_pnl,
                        "closed_at_raw": closed_at_raw,
                        "closed_at_str": closed_at_str,
                    })
            except Exception as e:
                logger.error(f"Error fetching analytics for account {account.name}: {e}")

        # If no closed trades, return empty
        if not all_closed:
            return jsonify({
                "status": "success",
                "metrics": {
                    "win_rate": 0,
                    "profit_factor": 0,
                    "sharpe_ratio": 0,
                    "recovery_factor": 0,
                    "total_trades": 0,
                    "net_profit": 0
                },
                "series_pnl": [],
                "series_drawdown": [],
                "heatmap": []
            })

        # Sort chronologically (oldest first) to build cumulative equity curve
        all_closed.sort(key=lambda x: x.get("closed_at_raw", 0))

        # 1. Math Analytics
        total_trades = len(all_closed)
        winning_trades = sum(1 for t in all_closed if t["net_pnl"] > 0)
        win_rate = (winning_trades / total_trades) * 100.0 if total_trades > 0 else 0.0
        
        gross_profit = sum(t["net_pnl"] for t in all_closed if t["net_pnl"] > 0)
        gross_loss = sum(abs(t["net_pnl"]) for t in all_closed if t["net_pnl"] < 0)
        profit_factor = (gross_profit / gross_loss) if gross_loss > 0 else (gross_profit if gross_profit > 0 else 1.0)
        net_profit = sum(t["net_pnl"] for t in all_closed)

        # Sharpe Ratio
        pnls = [t["net_pnl"] for t in all_closed]
        avg_pnl = sum(pnls) / total_trades
        if total_trades > 1:
            variance = sum((p - avg_pnl) ** 2 for p in pnls) / (total_trades - 1)
            std_dev = math.sqrt(variance)
            sharpe_ratio = (avg_pnl / std_dev) if std_dev > 0 else 0.0
        else:
            sharpe_ratio = 0.0

        # Equity Curve and Drawdowns
        running_equity = 0.0
        peak_equity = 0.0
        max_drawdown_pct = 0.0
        max_drawdown_usd = 0.0
        
        series_pnl = []
        series_drawdown = []
        
        # Add initial starting point
        series_pnl.append({"x": "Start", "y": 0.0})
        series_drawdown.append({"x": "Start", "y": 0.0})

        for idx, t in enumerate(all_closed):
            running_equity += t["net_pnl"]
            if running_equity > peak_equity:
                peak_equity = running_equity
            
            # Drawdown from peak in USD
            dd_usd = peak_equity - running_equity
            if dd_usd > max_drawdown_usd:
                max_drawdown_usd = dd_usd
                
            # Drawdown from peak in % (for the chart)
            dd_pct = (dd_usd / peak_equity * 100.0) if peak_equity > 0 else (dd_usd if dd_usd > 0 else 0.0)
            if dd_pct > max_drawdown_pct:
                max_drawdown_pct = dd_pct
                
            label = t["closed_at_str"] or f"Trade #{idx + 1}"
            series_pnl.append({"x": label, "y": round(running_equity, 4)})
            series_drawdown.append({"x": label, "y": round(-abs(dd_pct), 2)})

        recovery_factor = (net_profit / max_drawdown_usd) if max_drawdown_usd > 0 else 0.0

        # 2. Time-of-Day Heatmap Matrix
        heatmap_data = {day: {hour: 0.0 for hour in range(24)} for day in range(7)}
        
        for t in all_closed:
            if t["closed_at_raw"] > 0:
                import datetime
                dt = datetime.datetime.fromtimestamp(t["closed_at_raw"], datetime.timezone.utc)
                heatmap_data[dt.weekday()][dt.hour] += t["net_pnl"]

        day_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        heatmap_series = []
        
        for day_idx in range(7):
            day_name = day_names[day_idx]
            hour_data = []
            for hour in range(24):
                hour_label = f"{hour:02d}:00"
                hour_data.append({
                    "x": hour_label,
                    "y": round(heatmap_data[day_idx][hour], 4)
                })
            heatmap_series.append({
                "name": day_name,
                "data": hour_data
            })

        return jsonify({
            "status": "success",
            "metrics": {
                "win_rate": round(win_rate, 2),
                "profit_factor": round(profit_factor, 2),
                "sharpe_ratio": round(sharpe_ratio, 2),
                "recovery_factor": round(recovery_factor, 2),
                "total_trades": total_trades,
                "net_profit": round(net_profit, 4)
            },
            "series_pnl": series_pnl,
            "series_drawdown": series_drawdown,
            "heatmap": heatmap_series
        })
    except Exception as e:
        logger.exception(f"Error compiling analytics: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/logs", methods=["GET"])
def get_logs():
    logs = TradeLog.query.order_by(TradeLog.timestamp.desc()).limit(100).all()
    return jsonify([log.to_dict() for log in logs])

@app.route("/api/pnl", methods=["GET"])
def get_pnl():
    """Aggregates live open positions and closed trade history across all active accounts."""
    import datetime
    try:
        # Fetch active accounts
        active_accounts = Account.query.filter_by(is_active=True).all()
        if not active_accounts:
            if Config.API_KEY and Config.API_SECRET:
                fallback_account = Account(
                    id=0,
                    name="Environment Default",
                    api_key=Config.API_KEY,
                    api_secret=Config.API_SECRET,
                    leverage=Config.DEFAULT_LEVERAGE,
                    balance_buffer_pct=Config.BALANCE_BUFFER_PCT * 100.0,
                    sizing_type="percentage",
                    fixed_amount=10.0,
                    is_active=True
                )
                active_accounts = [fallback_account]
                
        # Fetch and map products for symbol translation
        products = []
        try:
            products = public_delta_client.get_products()
        except Exception as e:
            logger.warning(f"Failed to fetch products for symbol translation in /api/pnl: {e}")
            
        product_map = {p.get("id"): p for p in products if p.get("id")}
        
        open_positions = []
        closed_positions = []
        
        for account in active_accounts:
            # Open Positions
            try:
                client = DeltaClient(
                    api_key=account.api_key,
                    api_secret=account.api_secret,
                    base_url=Config.BASE_URL
                )
                positions = client.get_open_positions()
                for pos in positions:
                    size_val = float(pos.get("size") or 0)
                    if size_val == 0:
                        continue
                        
                    product_id = pos.get("product_id")
                    prod_info = product_map.get(product_id) or pos.get("product") or {}
                    symbol = prod_info.get("symbol") or f"ID:{product_id}"
                    
                    side_raw = pos.get("side", "").lower()
                    if side_raw in ["buy", "long"]:
                        side = "LONG"
                    elif side_raw in ["sell", "short"]:
                        side = "SHORT"
                    else:
                        side = "LONG" if size_val > 0 else "SHORT"
                        
                    upnl = pos.get("unrealized_pnl")
                    if upnl is None:
                        upnl = pos.get("upnl")
                    if upnl is None:
                        upnl = pos.get("pnl")
                    if upnl is None:
                        upnl = 0.0
                        
                    open_positions.append({
                        "account_name": account.name,
                        "product_id": product_id,
                        "symbol": symbol,
                        "side": side,
                        "size": abs(size_val),
                        "entry_price": float(pos.get("entry_price") or pos.get("avg_entry_price") or 0),
                        "mark_price": float(pos.get("mark_price") or 0),
                        "unrealized_pnl": float(upnl),
                        "margin": float(pos.get("margin") or 0),
                        "leverage": pos.get("leverage") or account.leverage
                    })
            except Exception as e:
                logger.exception(f"Error fetching open positions for account {account.name}: {e}")
                
            # Closed Positions
            try:
                client = DeltaClient(
                    api_key=account.api_key,
                    api_secret=account.api_secret,
                    base_url=Config.BASE_URL
                )
                closed = client.get_closed_positions(limit=150)
                for pos in closed:
                    product_id = pos.get("product_id")
                    prod_info = product_map.get(product_id) or pos.get("product") or {}
                    symbol = prod_info.get("symbol") or f"ID:{product_id}"
                    
                    side_raw = pos.get("side", "").lower()
                    if side_raw in ["buy", "long"]:
                        side = "LONG"
                    elif side_raw in ["sell", "short"]:
                        side = "SHORT"
                    else:
                        side = "LONG"
                        
                    closed_size = pos.get("closed_size")
                    if closed_size is None:
                        closed_size = pos.get("size")
                    if closed_size is None:
                        closed_size = 0.0
                        
                    rpnl = pos.get("realized_pnl")
                    if rpnl is None:
                        rpnl = pos.get("rpnl")
                    if rpnl is None:
                        rpnl = pos.get("pnl")
                    if rpnl is None:
                        rpnl = 0.0
                        
                    closed_at = pos.get("closed_at")
                    closed_at_str = ""
                    closed_at_raw = 0
                    if closed_at:
                        try:
                            iso_str = str(closed_at)
                            if iso_str.endswith('Z'):
                                iso_str = iso_str[:-1] + '+00:00'
                            try:
                                dt = datetime.datetime.fromisoformat(iso_str)
                                closed_at_raw = dt.timestamp()
                                closed_at_str = dt.strftime("%Y-%m-%d %H:%M:%S UTC")
                            except ValueError:
                                t_val = float(closed_at)
                                if t_val > 1e12:
                                    t_val = t_val / 1000.0
                                if t_val > 1e11:
                                    t_val = t_val / 1000.0
                                closed_at_raw = t_val
                                dt = datetime.datetime.fromtimestamp(t_val, datetime.timezone.utc)
                                closed_at_str = dt.strftime("%Y-%m-%d %H:%M:%S UTC")
                        except Exception:
                            closed_at_str = str(closed_at)
                            
                    entry_at = pos.get("entry_at")
                    entry_at_str = ""
                    entry_at_raw = 0
                    if entry_at:
                        try:
                            iso_str = str(entry_at)
                            if iso_str.endswith('Z'):
                                iso_str = iso_str[:-1] + '+00:00'
                            try:
                                dt = datetime.datetime.fromisoformat(iso_str)
                                entry_at_raw = dt.timestamp()
                                entry_at_str = dt.strftime("%Y-%m-%d %H:%M:%S UTC")
                            except ValueError:
                                t_val = float(entry_at)
                                if t_val > 1e12:
                                    t_val = t_val / 1000.0
                                if t_val > 1e11:
                                    t_val = t_val / 1000.0
                                entry_at_raw = t_val
                                dt = datetime.datetime.fromtimestamp(t_val, datetime.timezone.utc)
                                entry_at_str = dt.strftime("%Y-%m-%d %H:%M:%S UTC")
                        except Exception:
                            entry_at_str = str(entry_at)
                            
                    hold_time = max(0.0, closed_at_raw - entry_at_raw) if (entry_at_raw > 0 and closed_at_raw > 0) else None

                    # Determine contract value from product specifications
                    contract_val_str = prod_info.get("contract_value") or "0.01"
                    try:
                        contract_value = float(contract_val_str)
                    except ValueError:
                        contract_value = 0.01

                    # Retrieve actual commission fee or calculate 0.05% taker fee round-trip fallback
                    realized_fee_val = pos.get("fee") or pos.get("realized_fee") or pos.get("commission")
                    if realized_fee_val is not None:
                        try:
                            fees = abs(float(realized_fee_val))
                        except (ValueError, TypeError):
                            fees = 0.0
                    else:
                        entry_px = float(pos.get("entry_price") or pos.get("avg_entry_price") or 0)
                        exit_px = float(pos.get("close_price") or pos.get("exit_price") or pos.get("avg_exit_price") or 0)
                        c_size = abs(float(closed_size))
                        entry_notional = entry_px * c_size * contract_value
                        exit_notional = exit_px * c_size * contract_value
                        fees = (entry_notional + exit_notional) * 0.0005

                    net_pnl = float(rpnl) - fees

                    closed_positions.append({
                        "account_id": account.id,
                        "account_name": account.name,
                        "product_id": product_id,
                        "symbol": symbol,
                        "side": side,
                        "closed_size": abs(float(closed_size)),
                        "entry_price": float(pos.get("entry_price") or pos.get("avg_entry_price") or 0),
                        "close_price": float(pos.get("close_price") or pos.get("exit_price") or pos.get("avg_exit_price") or 0),
                        "realized_pnl": float(rpnl),
                        "fees": fees,
                        "net_pnl": net_pnl,
                        "closed_at": closed_at_str,
                        "closed_at_raw": closed_at_raw,
                        "entry_at": entry_at_str,
                        "hold_time": hold_time
                    })
            except Exception as e:
                logger.exception(f"Error fetching closed positions for account {account.name}: {e}")
                
        # Sort closed positions by timestamp descending
        closed_positions.sort(key=lambda x: x.get("closed_at_raw", 0), reverse=True)
        
        # Calculate Advanced Risk & Performance Metrics
        net_pnls = [t["net_pnl"] for t in closed_positions]
        total_trades = len(closed_positions)
        
        win_rate = 0.0
        profit_factor = 0.0
        sharpe_ratio = 0.0
        max_drawdown = 0.0
        avg_hold_time_val = 0.0
        avg_hold_time_str = "N/A"
        
        if total_trades > 0:
            # Win Rate
            wins = [p for p in net_pnls if p > 0]
            win_rate = (len(wins) / total_trades) * 100.0
            
            # Profit Factor
            gross_profits = sum(wins)
            gross_losses = abs(sum([p for p in net_pnls if p < 0]))
            if gross_losses > 0:
                profit_factor = gross_profits / gross_losses
            else:
                profit_factor = gross_profits if gross_profits > 0 else 0.0
                
            # Sharpe Ratio
            import math
            mean_pnl = sum(net_pnls) / total_trades
            if total_trades > 1:
                variance = sum((x - mean_pnl) ** 2 for x in net_pnls) / (total_trades - 1)
                std_pnl = math.sqrt(variance)
                sharpe_ratio = mean_pnl / std_pnl if std_pnl > 0 else 0.0
            else:
                sharpe_ratio = 0.0
                
            # Max Drawdown
            equity = 0.0
            equity_curve = [0.0]
            for t in reversed(closed_positions):
                equity += t["net_pnl"]
                equity_curve.append(equity)
            
            peak = equity_curve[0]
            for val in equity_curve:
                if val > peak:
                    peak = val
                dd = peak - val
                if dd > max_drawdown:
                    max_drawdown = dd
                    
            # Average Hold Time
            hold_times = [t["hold_time"] for t in closed_positions if t.get("hold_time") is not None]
            if hold_times:
                avg_hold_time_val = sum(hold_times) / len(hold_times)
                
                def format_duration(seconds):
                    if seconds <= 0:
                        return "0s"
                    parts = []
                    days = int(seconds // 86400)
                    if days > 0:
                        parts.append(f"{days}d")
                        seconds %= 86400
                    hours = int(seconds // 3600)
                    if hours > 0:
                        parts.append(f"{hours}h")
                        seconds %= 3600
                    minutes = int(seconds // 60)
                    if minutes > 0:
                        parts.append(f"{minutes}m")
                        seconds %= 60
                    secs = int(seconds)
                    if secs > 0 or not parts:
                        parts.append(f"{secs}s")
                    return " ".join(parts)
                
                avg_hold_time_str = format_duration(avg_hold_time_val)
        
        metrics = {
            "win_rate": round(win_rate, 2),
            "profit_factor": round(profit_factor, 2),
            "sharpe_ratio": round(sharpe_ratio, 2),
            "max_drawdown": round(max_drawdown, 2),
            "avg_hold_time": round(avg_hold_time_val, 2),
            "avg_hold_time_str": avg_hold_time_str,
            "total_trades": total_trades
        }

        # Remove closed_at_raw from response payload
        for item in closed_positions:
            item.pop("closed_at_raw", None)
            
        return jsonify({
            "open": open_positions,
            "closed": closed_positions,
            "metrics": metrics
        })
    except Exception as e:
        logger.exception(f"Error in /api/pnl endpoint: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/ai-insights", methods=["GET"])
def get_ai_insights():
    try:
        insights = AIInsight.query.order_by(AIInsight.timestamp.desc()).all()
        return jsonify([i.to_dict() for i in insights])
    except Exception as e:
        logger.exception(f"Error in GET /api/ai-insights: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/ai-insights/generate", methods=["POST"])
def generate_ai_insights():
    try:
        # 1. Retrieve the AI settings from database
        settings_keys = ["ai_api_key", "ai_provider", "ai_model"]
        settings = GlobalSetting.query.filter(GlobalSetting.key.in_(settings_keys)).all()
        s_dict = {s.key: s.value for s in settings}
        
        api_key = s_dict.get("ai_api_key")
        provider = s_dict.get("ai_provider", "gemini").lower()
        model = s_dict.get("ai_model", "gemini-1.5-flash")
        
        # Mismatch resolution logic
        if provider == "openai" and ("gemini" in model.lower() or not model):
            model = "gpt-4o-mini"
        elif provider == "gemini" and ("gpt" in model.lower() or not model):
            model = "gemini-1.5-flash"
        
        if not api_key:
            return jsonify({"status": "error", "message": "AI API Key is missing. Please configure it in Settings."}), 400
            
        # 2. Gather closed trade history from active accounts (up to 50 trades for analysis)
        active_accounts = Account.query.filter_by(is_active=True).all()
        closed_positions = []
        
        for account in active_accounts:
            try:
                client = DeltaClient(
                    api_key=account.api_key,
                    api_secret=account.api_secret,
                    base_url=Config.BASE_URL
                )
                closed = client.get_closed_positions(limit=50)
                for pos in closed:
                    symbol = pos.get("symbol") or "ETHUSD"
                    contract_value = 0.01
                    
                    is_india = "india" in Config.BASE_URL.lower()
                    
                    rpnl = float(pos.get("realized_pnl") or 0.0)
                    closed_size = float(pos.get("closed_size") or pos.get("size") or 0.0)
                    if closed_size == 0:
                        continue
                        
                    if is_india:
                        rpnl = rpnl / 85.0
                        
                    realized_fee_val = pos.get("fee")
                    if realized_fee_val is not None:
                        try:
                            fees = abs(float(realized_fee_val))
                            if is_india:
                                fees = fees / 85.0
                        except (ValueError, TypeError):
                            fees = 0.0
                    else:
                        entry_px = float(pos.get("entry_price") or pos.get("avg_entry_price") or 0)
                        exit_px = float(pos.get("close_price") or pos.get("exit_price") or pos.get("avg_exit_price") or 0)
                        c_size = abs(float(closed_size))
                        entry_notional = entry_px * c_size * contract_value
                        exit_notional = exit_px * c_size * contract_value
                        fees = (entry_notional + exit_notional) * 0.0005
                        
                    net_pnl = rpnl - fees
                    closed_at_str = pos.get("closed_at") or pos.get("created_at") or ""
                    
                    closed_positions.append({
                        "account_name": account.name,
                        "symbol": symbol,
                        "side": pos.get("side"),
                        "closed_size": closed_size,
                        "entry_price": float(pos.get("entry_price") or pos.get("avg_entry_price") or 0),
                        "close_price": float(pos.get("close_price") or pos.get("exit_price") or pos.get("avg_exit_price") or 0),
                        "realized_pnl": rpnl,
                        "fees": fees,
                        "net_pnl": net_pnl,
                        "closed_at": closed_at_str,
                    })
            except Exception as ex:
                logger.error(f"Error fetching trades for account {account.name}: {ex}")
                
        if not closed_positions:
            return jsonify({"status": "error", "message": "No trade history found to analyze."}), 400
            
        # Sort positions descending by date
        closed_positions.sort(key=lambda x: x.get("closed_at", ""), reverse=True)
        analysis_trades = closed_positions[:50]
        
        # Calculate summary stats
        net_pnls = [t["net_pnl"] for t in analysis_trades]
        total_trades = len(analysis_trades)
        wins = [p for p in net_pnls if p > 0]
        win_rate = (len(wins) / total_trades) * 100.0 if total_trades > 0 else 0.0
        
        gross_profits = sum(wins)
        gross_losses = abs(sum([p for p in net_pnls if p < 0]))
        profit_factor = gross_profits / gross_losses if gross_losses > 0 else (gross_profits if gross_profits > 0 else 0.0)
        
        # Compile trade history snippet for LLM prompt
        trades_text = ""
        for i, t in enumerate(analysis_trades, start=1):
            trades_text += f"{i}. Account: {t['account_name']} | Ticker: {t['symbol']} | Side: {t['side']} | Size: {t['closed_size']} | Entry: {t['entry_price']} | Close: {t['close_price']} | Realized PnL: ${t['realized_pnl']:.2f} | Fees: ${t['fees']:.2f} | Net PnL: ${t['net_pnl']:.2f} | Closed At: {t['closed_at']}\n"
            
        prompt = f"""You are an elite quantitative trading system analyst and risk consultant.
Review the following trading performance data and closed trade logs for our algorithmic bot.
Identify performance leaks, strategy mistakes, execution anomalies, and suggest concrete fixes for either the TradingView Pine Script strategy or the bot's risk configuration.

### Summary Metrics (Last {total_trades} Trades):
- Win Rate: {win_rate:.2f}%
- Profit Factor: {profit_factor:.2f}
- Total Net PnL: ${sum(net_pnls):.2f}
- Gross Profits: ${gross_profits:.2f}
- Gross Losses: ${gross_losses:.2f}

### Closed Trade Logs (Latest {total_trades} Trades):
{trades_text}

### Instructions for your Analysis:
Provide a highly professional, structured report in markdown:
1. **Executive Summary**: High-level review of performance and overall behavior.
2. **Key Performance Leakages**:
   - Identify if fees are consuming too much profit (fee drag).
   - Check if the strategy is cutting winners short and holding losers too long (average gain vs loss).
   - Check for over-trading or clusters of losing trades in tight windows.
   - Detect size/leverage anomalies.
3. **Actionable Recommendations**: Specific changes to variables (e.g. adjust leverage, buffer size, stop-loss ratios).
4. **Strategy Pine Script Adjustments**: Suggestions on what indicators or filters (like volume, ATR, session timings) to add in the TradingView strategy to filter noise.

Be direct, critical, and constructive. Highlight specific trade numbers to prove your points.
"""

        generated_text = ""
        import json
        
        if provider == "gemini":
            url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}"
            payload = {
                "contents": [{
                    "parts": [{
                        "text": prompt
                    }]
                }]
            }
            res = requests.post(url, json=payload, timeout=30)
            if res.status_code == 200:
                res_data = res.json()
                try:
                    generated_text = res_data['candidates'][0]['content']['parts'][0]['text']
                except KeyError:
                    logger.error(f"Unexpected Gemini response structure: {res_data}")
                    return jsonify({"status": "error", "message": "Failed to parse Gemini response."}), 500
            else:
                logger.error(f"Gemini API error: {res.status_code} - {res.text}")
                return jsonify({"status": "error", "message": f"Gemini API returned status code {res.status_code}: {res.text}"}), 500
                
        elif provider == "openai":
            url = "https://api.openai.com/v1/chat/completions"
            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json"
            }
            payload = {
                "model": model if model else "gpt-4o-mini",
                "messages": [
                    {"role": "user", "content": prompt}
                ]
            }
            res = requests.post(url, headers=headers, json=payload, timeout=30)
            if res.status_code == 200:
                res_data = res.json()
                try:
                    generated_text = res_data['choices'][0]['message']['content']
                except KeyError:
                    logger.error(f"Unexpected OpenAI response structure: {res_data}")
                    return jsonify({"status": "error", "message": "Failed to parse OpenAI response."}), 500
            else:
                logger.error(f"OpenAI API error: {res.status_code} - {res.text}")
                return jsonify({"status": "error", "message": f"OpenAI API returned status code {res.status_code}: {res.text}"}), 500
        else:
            return jsonify({"status": "error", "message": f"Unsupported AI provider: {provider}"}), 400
            
        if not generated_text:
            return jsonify({"status": "error", "message": "LLM generated an empty response."}), 500
            
        metrics_snapshot = {
            "win_rate": round(win_rate, 2),
            "profit_factor": round(profit_factor, 2),
            "total_net_pnl": round(sum(net_pnls), 2),
            "total_trades": total_trades
        }
        insight = AIInsight(
            insights=generated_text,
            metrics_snapshot=json.dumps(metrics_snapshot)
        )
        db.session.add(insight)
        db.session.commit()
        
        return jsonify(insight.to_dict())
        
    except Exception as e:
        logger.exception(f"Error in /api/ai-insights/generate endpoint: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/journal/export", methods=["GET"])
def export_journal():
    """Generates and downloads a CSV trade journal aggregating all closed positions across active accounts."""
    import csv
    import io
    import datetime
    try:
        # Retrieve optional account_id query parameter
        account_id_param = request.args.get("account_id")
        
        active_accounts = []
        if account_id_param and account_id_param.lower() != "all":
            try:
                acc_id = int(account_id_param)
                if acc_id == 0:
                    if Config.API_KEY and Config.API_SECRET:
                        fallback_account = Account(
                            id=0,
                            name="Environment Default",
                            api_key=Config.API_KEY,
                            api_secret=Config.API_SECRET,
                            leverage=Config.DEFAULT_LEVERAGE,
                            balance_buffer_pct=Config.BALANCE_BUFFER_PCT * 100.0,
                            sizing_type="percentage",
                            fixed_amount=10.0,
                            is_active=True
                        )
                        active_accounts = [fallback_account]
                else:
                    target_acc = Account.query.get(acc_id)
                    if target_acc:
                        active_accounts = [target_acc]
            except ValueError:
                pass
        else:
            active_accounts = Account.query.filter_by(is_active=True).all()
            if not active_accounts:
                if Config.API_KEY and Config.API_SECRET:
                    fallback_account = Account(
                        id=0,
                        name="Environment Default",
                        api_key=Config.API_KEY,
                        api_secret=Config.API_SECRET,
                        leverage=Config.DEFAULT_LEVERAGE,
                        balance_buffer_pct=Config.BALANCE_BUFFER_PCT * 100.0,
                        sizing_type="percentage",
                        fixed_amount=10.0,
                        is_active=True
                    )
                    active_accounts = [fallback_account]

        products = []
        try:
            products = public_delta_client.get_products()
        except Exception as e:
            logger.warning(f"Failed to fetch products for symbol translation in /api/journal/export: {e}")
            
        product_map = {p.get("id"): p for p in products if p.get("id")}
        
        closed_positions = []
        
        for account in active_accounts:
            try:
                client = DeltaClient(
                    api_key=account.api_key,
                    api_secret=account.api_secret,
                    base_url=Config.BASE_URL
                )
                closed = client.get_closed_positions(limit=100)
                for pos in closed:
                    product_id = pos.get("product_id")
                    prod_info = product_map.get(product_id) or pos.get("product") or {}
                    symbol = prod_info.get("symbol") or f"ID:{product_id}"
                    
                    side_raw = pos.get("side", "").lower()
                    if side_raw in ["buy", "long"]:
                        side = "LONG"
                    elif side_raw in ["sell", "short"]:
                        side = "SHORT"
                    else:
                        side = "LONG"
                        
                    closed_size = pos.get("closed_size")
                    if closed_size is None:
                        closed_size = pos.get("size")
                    if closed_size is None:
                        closed_size = 0.0
                        
                    rpnl = pos.get("realized_pnl")
                    if rpnl is None:
                        rpnl = pos.get("rpnl")
                    if rpnl is None:
                        rpnl = pos.get("pnl")
                    if rpnl is None:
                        rpnl = 0.0
                        
                    closed_at = pos.get("closed_at")
                    closed_at_str = ""
                    closed_at_raw = 0
                    if closed_at:
                        try:
                            iso_str = str(closed_at)
                            if iso_str.endswith('Z'):
                                iso_str = iso_str[:-1] + '+00:00'
                            try:
                                dt = datetime.datetime.fromisoformat(iso_str)
                                closed_at_raw = dt.timestamp()
                                closed_at_str = dt.strftime("%Y-%m-%d %H:%M:%S UTC")
                            except ValueError:
                                t_val = float(closed_at)
                                if t_val > 1e12:
                                    t_val = t_val / 1000.0
                                if t_val > 1e11:
                                    t_val = t_val / 1000.0
                                closed_at_raw = t_val
                                dt = datetime.datetime.fromtimestamp(t_val, datetime.timezone.utc)
                                closed_at_str = dt.strftime("%Y-%m-%d %H:%M:%S UTC")
                        except Exception:
                            closed_at_str = str(closed_at)

                    # Determine contract value
                    contract_val_str = prod_info.get("contract_value") or "0.01"
                    try:
                        contract_value = float(contract_val_str)
                    except ValueError:
                        contract_value = 0.01

                    # Retrieve actual commission fee or calculate 0.05% taker fee round-trip fallback
                    realized_fee_val = pos.get("fee") or pos.get("realized_fee") or pos.get("commission")
                    if realized_fee_val is not None:
                        try:
                            fees = abs(float(realized_fee_val))
                        except (ValueError, TypeError):
                            fees = 0.0
                    else:
                        entry_px = float(pos.get("entry_price") or pos.get("avg_entry_price") or 0)
                        exit_px = float(pos.get("close_price") or pos.get("exit_price") or pos.get("avg_exit_price") or 0)
                        c_size = abs(float(closed_size))
                        entry_notional = entry_px * c_size * contract_value
                        exit_notional = exit_px * c_size * contract_value
                        fees = (entry_notional + exit_notional) * 0.0005

                    net_pnl = float(rpnl) - fees
                    
                    closed_positions.append({
                        "closed_at": closed_at_str,
                        "closed_at_raw": closed_at_raw,
                        "account_name": account.name,
                        "symbol": symbol,
                        "side": side,
                        "closed_size": abs(float(closed_size)),
                        "entry_price": float(pos.get("entry_price") or pos.get("avg_entry_price") or 0),
                        "close_price": float(pos.get("close_price") or pos.get("exit_price") or pos.get("avg_exit_price") or 0),
                        "realized_pnl": float(rpnl),
                        "fees": fees,
                        "net_pnl": net_pnl
                    })
            except Exception as e:
                logger.exception(f"Error fetching closed positions for CSV export on account {account.name}: {e}")
                
        # Sort chronologically by timestamp descending
        closed_positions.sort(key=lambda x: x.get("closed_at_raw", 0), reverse=True)
        
        # Build Excel file
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trading Journal"
        
        # Enable grid lines explicitly
        ws.views.sheetView[0].showGridLines = True
        
        # Exact column widths matching target template
        column_widths = {
            'A': 22.0, 'B': 14.0, 'C': 10.0, 'D': 8.0, 'E': 20.0,
            'F': 18.0, 'G': 17.0, 'H': 16.0, 'I': 22.0, 'J': 14.0
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Define Styles
        font_header = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
        fill_header = PatternFill(fill_type="solid", fgColor="FF1A1A2E")
        align_center = Alignment(horizontal="center", vertical="center")
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        border_thin_side = Side(style="thin", color="FFCCCCCC")
        border_thin = Border(left=border_thin_side, right=border_thin_side, top=border_thin_side, bottom=border_thin_side)
        
        border_medium_side = Side(style="medium", color="FF000000")
        border_medium = Border(left=border_medium_side, right=border_medium_side, top=border_medium_side, bottom=border_medium_side)

        # Set header row height to 32.0 to ensure text wraps fully and is readable
        ws.row_dimensions[1].height = 32.0

        # Headers
        headers = [
            "CLOSED TIME (UTC)", "ACCOUNT NAME", "SYMBOL", "SIDE",
            "CLOSED SIZE (CONTRACTS)", "ENTRY PRICE (USD)", "EXIT PRICE (USD)",
            "GROSS PNL (USD)", "FEES & COMMISSION (USD)", "NET PNL (USD)"
        ]
        ws.append(headers)
        for col_idx in range(1, 11):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_header
            cell.border = border_medium

        for row_idx, pos in enumerate(closed_positions, start=2):
            is_win = pos["net_pnl"] >= 0
            fill_color = "FFE8F5E9" if is_win else "FFFFEBEE"
            text_color = "FF1B7E1B" if is_win else "FFC62828"
            
            fill_row = PatternFill(fill_type="solid", fgColor=fill_color)
            font_normal = Font(name="Arial", size=10, bold=False)
            font_pnl = Font(name="Arial", size=10, bold=True, color=text_color)
            font_fees = Font(name="Arial", size=10, bold=False, color="FFC62828")
            
            row_values = [
                pos["closed_at"], pos["account_name"], pos["symbol"], pos["side"],
                pos["closed_size"], pos["entry_price"], pos["close_price"],
                pos["realized_pnl"], -pos["fees"], pos["net_pnl"]
            ]
            
            ws.append(row_values)
            for col_idx in range(1, 11):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = fill_row
                cell.alignment = align_center
                cell.border = border_thin
                
                # Formatting
                if col_idx in [5, 6, 7, 8, 9, 10]:
                    cell.number_format = "#,##0.0000"
                else:
                    cell.number_format = "@"
                    
                if col_idx in [8, 10]: # Gross PNL and Net PNL are bold and colored
                    cell.font = font_pnl
                elif col_idx == 9: # Fees are colored red to represent loss
                    cell.font = font_fees
                else:
                    cell.font = font_normal

        # Total Row
        # Guard last_row to be at least 2 (if there are no trades) to avoid backward ranges like H2:H1
        last_row = max(len(closed_positions) + 1, 2)
        total_row = last_row + 1
        
        # Calculate sum dynamically to determine coloring and write values directly
        total_gross = sum(p["realized_pnl"] for p in closed_positions) if closed_positions else 0.0
        total_fees = sum(p["fees"] for p in closed_positions) if closed_positions else 0.0
        total_net = sum(p["net_pnl"] for p in closed_positions) if closed_positions else 0.0
        
        ws.cell(row=total_row, column=1, value="TOTAL")
        ws.cell(row=total_row, column=8, value=total_gross)
        ws.cell(row=total_row, column=9, value=-total_fees)
        ws.cell(row=total_row, column=10, value=total_net)
        
        color_gross = "FF1B7E1B" if total_gross >= 0 else "FFC62828"
        color_net = "FF1B7E1B" if total_net >= 0 else "FFC62828"
        
        fill_total = PatternFill(fill_type="solid", fgColor="FFE8EAF6")
        font_total_lbl = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
        font_total_val = Font(name="Calibri", size=11, bold=False)
        font_total_gross = Font(name="Arial", size=10, bold=True, color=color_gross)
        font_total_fees = Font(name="Arial", size=10, bold=True, color="FFC62828")
        font_total_net = Font(name="Arial", size=10, bold=True, color=color_net)

        for col_idx in range(1, 11):
            cell = ws.cell(row=total_row, column=col_idx)
            cell.alignment = align_center
            cell.border = border_medium
            
            if col_idx == 1:
                cell.font = font_total_lbl
                cell.fill = fill_header
            else:
                cell.fill = fill_total
                if col_idx == 8:
                    cell.font = font_total_gross
                    cell.number_format = "#,##0.0000"
                elif col_idx == 9:
                    cell.font = font_total_fees
                    cell.number_format = "#,##0.0000"
                elif col_idx == 10:
                    cell.font = font_total_net
                    cell.number_format = "#,##0.0000"
                else:
                    cell.font = font_total_val
                    
        # Force formula evaluation on load
        wb.calculation.calcMode = 'auto'
        wb.calculation.fullCalcOnLoad = True

        # Save to BytesIO buffer
        out_buf = io.BytesIO()
        wb.save(out_buf)
        out_buf.seek(0)
        
        # Determine output filename dynamically
        filename = "trading_journal.xlsx"
        if len(active_accounts) == 1:
            clean_name = active_accounts[0].name.lower().replace(" ", "_")
            filename = f"trading_journal_{clean_name}.xlsx"
            
        output = make_response(out_buf.getvalue())
        output.headers["Content-Disposition"] = f"attachment; filename={filename}"
        output.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return output
        
    except Exception as e:
        logger.exception(f"Error exporting trade journal: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/local-signals/export", methods=["GET"])
def export_local_signals():
    """Generates and downloads an Excel file containing all local strategy dry-run signal logs."""
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    try:
        # Fetch all signal logs sorted by timestamp descending
        logs = LocalSignalLog.query.order_by(LocalSignalLog.timestamp.desc()).all()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Local Bot Signals"
        
        ws.views.sheetView[0].showGridLines = True
        
        # Column widths
        column_widths = {
            'A': 22.0,  # Date/Time
            'B': 18.0,  # Account Name
            'C': 15.0,  # Signal Type
            'D': 15.0,  # Price
            'E': 15.0,  # Quantity (Lots)
            'F': 15.0,  # Stop Loss
            'G': 15.0,  # Take Profit 1
            'H': 15.0,  # Take Profit 2
            'I': 15.0   # Matched?
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
            
        # Styles
        font_header = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
        # Dark purple theme for local signals to distinguish from trade journal
        fill_header = PatternFill(fill_type="solid", fgColor="FF4C1D95") 
        align_center = Alignment(horizontal="center", vertical="center")
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        border_thin_side = Side(style="thin", color="FFCCCCCC")
        border_thin = Border(left=border_thin_side, right=border_thin_side, top=border_thin_side, bottom=border_thin_side)
        
        border_medium_side = Side(style="medium", color="FF000000")
        border_medium = Border(left=border_medium_side, right=border_medium_side, top=border_medium_side, bottom=border_medium_side)
        
        ws.row_dimensions[1].height = 32.0
        
        headers = [
            "DATE/TIME (UTC)", "ACCOUNT NAME", "SIGNAL TYPE", "PRICE (USD)",
            "QUANTITY (LOTS)", "STOP LOSS (USD)", "TAKE PROFIT 1 (USD)", "TAKE PROFIT 2 (USD)", "MATCHED?"
        ]
        ws.append(headers)
        for col_idx in range(1, 10):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_header
            cell.border = border_medium
            
        for row_idx, log in enumerate(logs, start=2):
            sig = log.signal_type.upper()
            
            # Highlight entries
            if sig in ["BUY", "LONG"]:
                fill_color = "FFE8F5E9"  # light green
                text_color = "FF1B7E1B"
            elif sig in ["SELL", "SHORT"]:
                fill_color = "FFFFEBEE"  # light red
                text_color = "FFC62828"
            elif sig in ["TP1", "TP2"]:
                fill_color = "FFE8EAF6"  # light blue
                text_color = "FF3F51B5"
            else:
                fill_color = "FFFFF9C4"  # light yellow
                text_color = "FFF57F17"
                
            fill_row = PatternFill(fill_type="solid", fgColor=fill_color)
            font_normal = Font(name="Arial", size=10, bold=False)
            font_bold_color = Font(name="Arial", size=10, bold=True, color=text_color)
            
            row_values = [
                log.timestamp.strftime("%Y-%m-%d %H:%M:%S") if log.timestamp else "",
                log.account_name,
                sig,
                log.price,
                log.quantity,
                log.stop_loss,
                log.take_profit_1,
                log.take_profit_2,
                "YES" if log.is_matched else "NO"
            ]
            
            ws.append(row_values)
            for col_idx in range(1, 10):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = fill_row
                cell.alignment = align_center
                cell.border = border_thin
                
                # Format numbers
                if col_idx in [4, 6, 7, 8]:
                    cell.number_format = "#,##0.0000"
                elif col_idx == 5:
                    cell.number_format = "#,##0"
                else:
                    cell.number_format = "@"
                    
                if col_idx == 3: # Signal Type is bold and colored
                    cell.font = font_bold_color
                else:
                    cell.font = font_normal
                    
        out_buf = io.BytesIO()
        wb.save(out_buf)
        out_buf.seek(0)
        
        response = make_response(out_buf.getvalue())
        response.headers["Content-Disposition"] = "attachment; filename=local_bot_signals.xlsx"
        response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return response
        
    except Exception as e:
        logger.exception(f"Failed to export local signals: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/email-settings", methods=["GET"])
def get_email_settings():
    def get_setting(key, default=""):
        s = GlobalSetting.query.filter_by(key=key).first()
        return s.value if s else default
        
    pwd = get_setting("email_password")
    masked_pwd = "********" if pwd else ""
    
    return jsonify({
        "enabled": get_setting("email_enabled", "false") == "true",
        "imap_host": get_setting("imap_host", "imap.gmail.com"),
        "imap_port": get_setting("imap_port", "993"),
        "email_address": get_setting("email_address"),
        "email_password": masked_pwd,
        "email_sender": get_setting("email_sender", "noreply@tradingview.com"),
        "email_subject": get_setting("email_subject", "TradingView Alert")
    })

@app.route("/api/email-settings", methods=["POST"])
def save_email_settings():
    data = request.get_json(silent=True) or {}
    
    def set_setting(key, val):
        if val is None:
            val = ""
        s = GlobalSetting.query.filter_by(key=key).first()
        if s:
            s.value = str(val)
        else:
            s = GlobalSetting(key=key, value=str(val))
            db.session.add(s)

    set_setting("email_enabled", "true" if data.get("enabled") else "false")
    set_setting("imap_host", data.get("imap_host"))
    set_setting("imap_port", data.get("imap_port", "993"))
    set_setting("email_address", data.get("email_address"))
    
    pwd = data.get("email_password")
    if pwd and pwd != "********":
        set_setting("email_password", pwd)
        
    set_setting("email_sender", data.get("email_sender", "noreply@tradingview.com"))
    set_setting("email_subject", data.get("email_subject", "TradingView Alert"))
    
    db.session.commit()
    return jsonify({"status": "success", "message": "Email settings saved"})

@app.route("/api/email-settings/test", methods=["POST"])
def test_email_connection():
    data = request.get_json(silent=True) or {}
    imap_host = data.get("imap_host")
    imap_port = data.get("imap_port", "993")
    email_address = data.get("email_address")
    email_password = data.get("email_password")
    
    if email_password == "********":
        pwd_setting = GlobalSetting.query.filter_by(key="email_password").first()
        email_password = pwd_setting.value if pwd_setting else ""
        
    if not imap_host or not email_address or not email_password:
        return jsonify({"status": "error", "message": "IMAP Host, Email Address, and Password are required"}), 400
        
    import imaplib
    try:
        port = int(imap_port)
        mail = imaplib.IMAP4_SSL(imap_host, port, timeout=10)
        mail.login(email_address, email_password)
        mail.logout()
        return jsonify({"status": "success", "message": "Connection and login successful!"})
    except Exception as e:
        logger.exception(f"IMAP test connection failed: {e}")
        return jsonify({"status": "error", "message": f"Connection failed: {str(e)}"}), 500

@app.route("/api/accounts", methods=["GET"])
def get_accounts():
    accounts = Account.query.all()
    return jsonify([acc.to_dict() for acc in accounts])

@app.route("/api/accounts", methods=["POST"])
def add_account():
    data = request.get_json(silent=True) or {}
    name = data.get("name")
    api_key = data.get("api_key")
    api_secret = data.get("api_secret")
    leverage = data.get("leverage", 50)
    balance_buffer_pct = data.get("balance_buffer_pct", 55.0)
    sizing_type = data.get("sizing_type", "percentage")
    fixed_amount = data.get("fixed_amount", 10.0)
    
    daily_loss_limit = None
    if "daily_loss_limit" in data and data["daily_loss_limit"] is not None and data["daily_loss_limit"] != "":
        try:
            daily_loss_limit = float(data["daily_loss_limit"])
        except ValueError:
            pass
            
    if not name or not api_key or not api_secret:
        return jsonify({"status": "error", "message": "Name, API Key, and API Secret are required"}), 400
        
    acc = Account(
        name=name,
        api_key=api_key,
        api_secret=api_secret,
        leverage=int(leverage),
        balance_buffer_pct=float(balance_buffer_pct),
        sizing_type=sizing_type,
        fixed_amount=float(fixed_amount),
        daily_loss_limit=daily_loss_limit,
        is_circuit_broken=False,
        is_active=True
    )
    db.session.add(acc)
    db.session.commit()
    return jsonify({"status": "success", "message": "Account added", "account": acc.to_dict()})

@app.route("/api/accounts/<int:id>", methods=["PUT"])
def update_account(id):
    acc = Account.query.get_or_404(id)
    data = request.get_json(silent=True) or {}
    if "is_active" in data:
        acc.is_active = bool(data["is_active"])
    if "name" in data:
        acc.name = data["name"]
    if "leverage" in data:
        acc.leverage = int(data["leverage"])
    if "balance_buffer_pct" in data:
        acc.balance_buffer_pct = float(data["balance_buffer_pct"])
    if "api_key" in data:
        acc.api_key = data["api_key"]
    if "api_secret" in data and data["api_secret"]:
        acc.api_secret = data["api_secret"]
    if "sizing_type" in data:
        acc.sizing_type = data["sizing_type"]
    if "fixed_amount" in data:
        acc.fixed_amount = float(data["fixed_amount"])
    if "daily_loss_limit" in data:
        limit_val = data["daily_loss_limit"]
        if limit_val is None or limit_val == "" or limit_val == "null":
            acc.daily_loss_limit = None
        else:
            try:
                acc.daily_loss_limit = float(limit_val)
            except ValueError:
                pass
    if "is_circuit_broken" in data:
        acc.is_circuit_broken = bool(data["is_circuit_broken"])
    if "local_strategy_enabled" in data:
        acc.local_strategy_enabled = bool(data["local_strategy_enabled"])
        
    db.session.commit()
    return jsonify({"status": "success", "account": acc.to_dict()})

@app.route("/api/accounts/<int:id>/reset-breaker", methods=["POST"])
def reset_breaker(id):
    acc = Account.query.get_or_404(id)
    acc.is_circuit_broken = False
    db.session.commit()
    logger.info(f"Circuit breaker reset successfully for account '{acc.name}' (ID: {acc.id}).")
    return jsonify({"status": "success", "message": f"Circuit breaker reset for account {acc.name}", "account": acc.to_dict()})

@app.route("/api/accounts/<int:id>/toggle-strategy", methods=["POST"])
def toggle_strategy(id):
    acc = Account.query.get_or_404(id)
    data = request.get_json(silent=True) or {}
    if "enabled" in data:
        acc.local_strategy_enabled = bool(data["enabled"])
    else:
        acc.local_strategy_enabled = not acc.local_strategy_enabled
    db.session.commit()
    logger.info(f"Local strategy enabled status set to {acc.local_strategy_enabled} for account '{acc.name}' (ID: {acc.id}).")
    return jsonify({
        "status": "success", 
        "message": f"Local strategy {'enabled' if acc.local_strategy_enabled else 'disabled'} for account {acc.name}", 
        "account": acc.to_dict()
    })

@app.route("/api/accounts/<int:id>", methods=["DELETE"])
def delete_account(id):
    acc = Account.query.get_or_404(id)
    db.session.delete(acc)
    db.session.commit()
    return jsonify({"status": "success", "message": "Account deleted"})

@app.route("/api/accounts/<int:id>/balance", methods=["GET"])
def get_account_balance(id):
    acc = Account.query.get_or_404(id)
    try:
        client = DeltaClient(
            api_key=acc.api_key,
            api_secret=acc.api_secret,
            base_url=Config.BASE_URL
        )
        balance, asset = client.get_available_balance()
        return jsonify({"success": True, "balance": balance, "asset": asset})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@app.route("/api/accounts/<int:account_id>/strategies", methods=["GET"])
def get_strategies(account_id):
    try:
        strategies = Strategy.query.filter_by(account_id=account_id).all()
        results = []
        for s in strategies:
            s_dict = s.to_dict()
            states = StrategyState.query.filter_by(account_id=account_id, strategy_id=s.id).all()
            s_dict["positions"] = []
            for state in states:
                if state.position_size != 0:
                    s_dict["positions"].append({
                        "symbol": state.symbol,
                        "position_size": state.position_size,
                        "entry_price": state.entry_price
                    })
            results.append(s_dict)
        return jsonify(results)
    except Exception as e:
        logger.exception(f"Error fetching strategies: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/accounts/<int:account_id>/strategies", methods=["POST"])
def add_strategy(account_id):
    try:
        data = request.get_json()
        if not data or not data.get("name"):
            return jsonify({"status": "error", "message": "Missing strategy name"}), 400
            
        # Verify account exists
        account = Account.query.get(account_id)
        if not account:
            return jsonify({"status": "error", "message": "Account not found"}), 404
            
        # Check max limit of 10 strategies
        existing_count = Strategy.query.filter_by(account_id=account_id).count()
        if existing_count >= 10:
            return jsonify({"status": "error", "message": "Maximum of 10 strategies limit reached for this account"}), 400
            
        # Check unique name per account
        name = data.get("name").strip()
        dup = Strategy.query.filter_by(account_id=account_id).filter(Strategy.name.ilike(name)).first()
        if dup:
            return jsonify({"status": "error", "message": f"A strategy named '{name}' already exists"}), 400
            
        def safe_float(val, default):
            if val is None:
                return default
            try:
                import math
                f_val = float(val)
                if math.isnan(f_val) or math.isinf(f_val):
                    return default
                return f_val
            except (ValueError, TypeError):
                return default

        strategy = Strategy(
            account_id=account_id,
            name=name,
            sizing_type=data.get("sizing_type", "percentage"),
            balance_buffer_pct=safe_float(data.get("balance_buffer_pct"), 10.0),
            fixed_amount=safe_float(data.get("fixed_amount"), 10.0),
            leverage=int(safe_float(data.get("leverage"), 50)),
            is_active=data.get("is_active", True)
        )
        db.session.add(strategy)
        db.session.commit()
        return jsonify({"status": "success", "strategy": strategy.to_dict()}), 201
    except Exception as e:
        db.session.rollback()
        logger.exception(f"Error adding strategy: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/strategies/<int:strategy_id>", methods=["PUT"])
def update_strategy(strategy_id):
    try:
        data = request.get_json()
        strategy = Strategy.query.get(strategy_id)
        if not strategy:
            return jsonify({"status": "error", "message": "Strategy not found"}), 404
            
        if "name" in data:
            name = data.get("name").strip()
            if not name:
                return jsonify({"status": "error", "message": "Strategy name cannot be empty"}), 400
            # Check unique name excluding self
            dup = Strategy.query.filter_by(account_id=strategy.account_id).filter(Strategy.name.ilike(name)).filter(Strategy.id != strategy_id).first()
            if dup:
                return jsonify({"status": "error", "message": f"A strategy named '{name}' already exists"}), 400
            strategy.name = name
            
        def safe_float(val, default):
            if val is None:
                return default
            try:
                import math
                f_val = float(val)
                if math.isnan(f_val) or math.isinf(f_val):
                    return default
                return f_val
            except (ValueError, TypeError):
                return default

        if "sizing_type" in data:
            strategy.sizing_type = data.get("sizing_type")
        if "balance_buffer_pct" in data:
            strategy.balance_buffer_pct = safe_float(data.get("balance_buffer_pct"), 10.0)
        if "fixed_amount" in data:
            strategy.fixed_amount = safe_float(data.get("fixed_amount"), 10.0)
        if "leverage" in data:
            strategy.leverage = int(safe_float(data.get("leverage"), 50))
            
        db.session.commit()
        return jsonify({"status": "success", "strategy": strategy.to_dict()})
    except Exception as e:
        db.session.rollback()
        logger.exception(f"Error updating strategy: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/strategies/<int:strategy_id>", methods=["DELETE"])
def delete_strategy(strategy_id):
    try:
        strategy = Strategy.query.get(strategy_id)
        if not strategy:
            return jsonify({"status": "error", "message": "Strategy not found"}), 404
            
        # Explicitly delete all associated StrategyState records
        StrategyState.query.filter_by(strategy_id=strategy_id).delete()
        
        db.session.delete(strategy)
        db.session.commit()
        return jsonify({"status": "success", "message": "Strategy deleted"})
    except Exception as e:
        db.session.rollback()
        logger.exception(f"Error deleting strategy: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/strategies/<int:strategy_id>/toggle", methods=["POST"])
def toggle_strategy_status(strategy_id):
    try:
        data = request.get_json() or {}
        strategy = Strategy.query.get(strategy_id)
        if not strategy:
            return jsonify({"status": "error", "message": "Strategy not found"}), 404
            
        is_active = data.get("is_active")
        if is_active is not None:
            strategy.is_active = bool(is_active)
        else:
            strategy.is_active = not strategy.is_active
            
        db.session.commit()
        return jsonify({"status": "success", "is_active": strategy.is_active})
    except Exception as e:
        db.session.rollback()
        logger.exception(f"Error toggling strategy status: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

# ----------------- TRADING WEBHOOK ENDPOINT -----------------

def get_daily_pnl(client, product_map=None):
    """Calculates cumulative net PnL (realized PnL - fees) for positions closed today in UTC."""
    import datetime
    now_utc = datetime.datetime.now(datetime.timezone.utc)
    start_of_day = datetime.datetime(now_utc.year, now_utc.month, now_utc.day, tzinfo=datetime.timezone.utc)
    start_timestamp = start_of_day.timestamp()
    
    if not product_map:
        try:
            products = public_delta_client.get_products()
            product_map = {p.get("id"): p for p in products if p.get("id")}
        except Exception:
            product_map = {}
            
    daily_net_pnl = 0.0
    try:
        closed = client.get_closed_positions(limit=50)
        for pos in closed:
            closed_at = pos.get("closed_at")
            closed_at_raw = 0
            if closed_at:
                try:
                    iso_str = str(closed_at)
                    if iso_str.endswith('Z'):
                        iso_str = iso_str[:-1] + '+00:00'
                    try:
                        dt = datetime.datetime.fromisoformat(iso_str)
                        closed_at_raw = dt.timestamp()
                    except ValueError:
                        t_val = float(closed_at)
                        if t_val > 1e12:
                            t_val = t_val / 1000.0
                        if t_val > 1e11:
                            t_val = t_val / 1000.0
                        closed_at_raw = t_val
                except Exception:
                    pass
            
            if closed_at_raw >= start_timestamp:
                rpnl = float(pos.get("realized_pnl") or pos.get("pnl") or 0.0)
                product_id = pos.get("product_id")
                prod_info = product_map.get(product_id) or pos.get("product") or {}
                
                realized_fee_val = pos.get("fee") or pos.get("realized_fee") or pos.get("commission")
                if realized_fee_val is not None:
                    fees = abs(float(realized_fee_val))
                else:
                    entry_px = float(pos.get("entry_price") or pos.get("avg_entry_price") or 0)
                    exit_px = float(pos.get("close_price") or pos.get("exit_price") or 0)
                    contract_val_str = prod_info.get("contract_value") or "0.01"
                    try:
                        contract_value = float(contract_val_str)
                    except ValueError:
                        contract_value = 0.01
                    c_size = abs(float(pos.get("closed_size") or pos.get("size") or 0.0))
                    entry_notional = entry_px * c_size * contract_value
                    exit_notional = exit_px * c_size * contract_value
                    fees = (entry_notional + exit_notional) * 0.0005
                
                net_pnl = rpnl - fees
                daily_net_pnl += net_pnl
    except Exception as e:
        logger.error(f"Error calculating daily PnL: {e}")
    return daily_net_pnl

def close_all_positions(client):
    """Closes all open positions using reduce-only orders and returns a list of summaries with PnLs."""
    results_summary = []
    try:
        positions = client.get_open_positions()
        for pos in positions:
            try:
                size_val = float(pos.get("size") or 0.0)
                pos_size = abs(int(size_val))
                if pos_size > 0:
                    product_id = pos.get("product_id")
                    symbol = pos.get("symbol")
                    if not symbol:
                        # Fallback symbol lookup if not in position
                        product = client.get_product_by_id(product_id)
                        symbol = product.get("symbol") if product else f"ID {product_id}"
                        
                    is_long = size_val > 0
                    entry_px = float(pos.get("entry_price") or pos.get("avg_entry_price") or 0.0)
                    
                    close_side = "sell" if is_long else "buy"
                    res = client.place_order(
                        product_id=product_id,
                        size=pos_size,
                        side=close_side,
                        order_type="market_order",
                        reduce_only=True
                    )
                    
                    pnl_val = 0.0
                    if res.get("success"):
                        order_res = res.get("result", {})
                        try:
                            exit_px = float(order_res.get("average_fill_price") or 0.0)
                        except (ValueError, TypeError):
                            exit_px = 0.0
                        
                        if exit_px <= 0:
                            ticker_data = client.get_ticker(symbol)
                            exit_px = float(ticker_data.get("mark_price") or ticker_data.get("last_price") or 0.0)
                            
                        if entry_px > 0 and exit_px > 0:
                            direction = 1 if is_long else -1
                            # Look up product details to get contract_value
                            contract_value = 0.01 # Default fallback
                            try:
                                product = client.get_product_by_symbol(symbol)
                                if product:
                                    contract_value = float(product.get("contract_value", "0.01"))
                            except Exception:
                                pass
                            pnl_val = (exit_px - entry_px) * pos_size * contract_value * direction
                        
                        pnl_str = f"PnL: {pnl_val:+.2f} USD"
                        results_summary.append(f"Closed {pos_size} lots of {symbol} ({pnl_str})")
                    else:
                        results_summary.append(f"Failed to close {symbol}: {res.get('message', 'API Error')}")
            except Exception as e:
                logger.error(f"Failed to close position in close_all_positions: {e}")
                results_summary.append(f"Error closing position {pos.get('symbol', product_id)}: {e}")
    except Exception as e:
        logger.error(f"Error in close_all_positions: {e}")
    return results_summary

def execute_trades_background(accounts_data, ticker, action, source="webhook", payload=None):
    """Processes trading signals across all configured accounts in a background thread."""
    # 1. Retrieve product details (network call, outside DB context)
    product = public_delta_client.get_product_by_symbol(ticker)
    if not product:
        logger.error(f"Symbol '{ticker}' not found on Delta Exchange in background. Aborting.")
        with app.app_context():
            log_entry = TradeLog(
                ticker=ticker,
                action=action,
                source=source,
                status="failed",
                details=f"Symbol '{ticker}' not found on Delta Exchange."
            )
            db.session.add(log_entry)
            db.session.commit()
            
        # Send Notification
        title = f"🔴 Trade Alert Failure: {ticker} ({action.upper()})"
        msg = f"Symbol <b>'{ticker}'</b> not found on Delta Exchange."
        send_notification(title, msg, 15680580, symbol=ticker)
        
        return [{
            "account_id": acc["id"],
            "name": acc["name"],
            "success": False,
            "message": f"Symbol {ticker} not found on Delta Exchange"
        } for acc in accounts_data]

    product_id = product.get("id")
    symbol = product.get("symbol")
    contract_value_str = product.get("contract_value", "0.01")
    try:
        contract_value = float(contract_value_str)
    except ValueError:
        contract_value = 0.01

    logger.info(f"Parsed Product in background: {symbol} (ID: {product_id}, Lot Size: {contract_value})")
    
    results = []
    
    for acc in accounts_data:
        acc_name = acc["name"]
        acc_id = acc["id"]
        account_result = {
            "account_id": acc_id,
            "name": acc_name,
            "success": False
        }
        trade_msgs = []
        logger.info(f"Background processing for account '{acc_name}' (ID: {acc_id})...")
        try:
            # Initialize account-specific Delta REST Client
            client = DeltaClient(
                api_key=acc["api_key"],
                api_secret=acc["api_secret"],
                base_url=Config.BASE_URL
            )
            
            # Retrieve fresh DB details in a short context
            is_circuit_broken_flag = False
            daily_loss_limit_val = None
            if acc_id != 0:
                with app.app_context():
                    acc_db = db.session.get(Account, acc_id)
                    if acc_db:
                        is_circuit_broken_flag = acc_db.is_circuit_broken
                        daily_loss_limit_val = acc_db.daily_loss_limit
            
            # Check circuit breaker before processing
            if is_circuit_broken_flag:
                if action in ["buy", "sell"]:
                    logger.warning(f"Trading halted on account '{acc_name}' (ID: {acc_id}): Daily drawdown circuit breaker is tripped.")
                    account_result.update({"success": False, "message": "Circuit breaker is broken (trading halted)"})
                    results.append(account_result)
                    continue
                    
            # Perform daily drawdown calculation if limit is configured (network call, outside DB context)
            if not is_circuit_broken_flag and daily_loss_limit_val is not None:
                daily_pnl = get_daily_pnl(client)
                logger.info(f"Account '{acc_name}' daily net PnL: {daily_pnl:.4f} USD (Limit: {daily_loss_limit_val:.4f} USD)")
                if daily_pnl < 0 and abs(daily_pnl) >= daily_loss_limit_val:
                    logger.warning(f"Daily loss limit reached on account '{acc_name}'! Breached limit: {daily_loss_limit_val:.2f} USD. Tripping circuit breaker...")
                    
                    # 1. Trip circuit breaker in DB (short context)
                    if acc_id != 0:
                        with app.app_context():
                            acc_db = db.session.get(Account, acc_id)
                            if acc_db:
                                acc_db.is_circuit_broken = True
                                db.session.commit()
                    
                    # 2. Close all positions (network call)
                    close_all_positions(client)
                    
                    # 3. Dispatch alert
                    title = f"🚨 Circuit Breaker Tripped: {acc_name}"
                    notification_message = (
                        f"Account: <b>{acc_name}</b>\n"
                        f"Status: <b>HALTED</b>\n"
                        f"Daily Loss Breach: <b>{abs(daily_pnl):.2f} USD</b> (Limit: {daily_loss_limit_val:.2f} USD)\n"
                        f"Action: <b>All positions automatically closed</b>"
                    )
                    send_notification(title, notification_message, 15549011)
                    
                    account_result.update({"success": False, "message": f"Circuit breaker tripped. Daily loss: {abs(daily_pnl):.2f} USD"})
                    results.append(account_result)
                    continue
            
            # Resolve Strategy and StrategyState (short context)
            strategy_name = None
            if payload:
                strategy_name = payload.get("strategy") or payload.get("strategy_name")
            
            strategy_id = None
            strategy_leverage = None
            strategy_sizing_type = None
            strategy_fixed_amount = None
            strategy_balance_buffer_pct = None
            strategy_is_active = True
            
            if strategy_name and acc_id != 0:
                strategy_name = strategy_name.strip()
                with app.app_context():
                    strat_db = Strategy.query.filter_by(account_id=acc_id).filter(Strategy.name.ilike(strategy_name)).first()
                    if strat_db:
                        strategy_id = strat_db.id
                        strategy_is_active = strat_db.is_active
                        strategy_leverage = strat_db.leverage
                        strategy_sizing_type = strat_db.sizing_type
                        strategy_fixed_amount = strat_db.fixed_amount
                        strategy_balance_buffer_pct = strat_db.balance_buffer_pct
                    else:
                        strategy_is_active = None
                
                if strategy_is_active is None:
                    logger.warning(f"Strategy '{strategy_name}' not found for account '{acc_name}'. Skipping execution.")
                    account_result.update({"success": False, "message": f"Strategy '{strategy_name}' not found"})
                    results.append(account_result)
                    continue
                if not strategy_is_active:
                    logger.info(f"Strategy '{strategy_name}' is inactive for account '{acc_name}'. Skipping execution.")
                    account_result.update({"success": True, "message": f"Strategy '{strategy_name}' is inactive, skipped"})
                    results.append(account_result)
                    continue
            
            # Fetch or create StrategyState for this virtual position (short context)
            state_id = None
            state_position_size = 0.0
            state_entry_price = None
            
            if acc_id != 0:
                with app.app_context():
                    state_db = StrategyState.query.filter_by(account_id=acc_id, symbol=symbol, strategy_id=strategy_id).first()
                    if not state_db:
                        state_db = StrategyState(account_id=acc_id, symbol=symbol, strategy_id=strategy_id, position_size=0.0)
                        db.session.add(state_db)
                        db.session.commit()
                    state_id = state_db.id
                    state_position_size = state_db.position_size
                    state_entry_price = state_db.entry_price

            # Check for close/exit actions
            if action in ["close_long", "close_short"]:
                if strategy_name:
                    # Strategy-specific close logic using virtual position size
                    pos_size = abs(int(float(state_position_size)))
                    if pos_size == 0:
                        logger.info(f"Virtual position size is 0 for strategy '{strategy_name}' on account '{acc_name}'.")
                        account_result.update({"success": True, "message": "No virtual position size to close"})
                        results.append(account_result)
                        continue

                    is_long = state_position_size > 0
                    is_short = state_position_size < 0

                    if action == "close_long" and not is_long:
                        logger.warning(f"Received close_long alert but virtual position is not LONG (size: {state_position_size}) for strategy '{strategy_name}' on account '{acc_name}'. Ignoring.")
                        account_result.update({"success": True, "message": "Current virtual position is not LONG, ignoring close_long"})
                        results.append(account_result)
                        continue

                    if action == "close_short" and not is_short:
                        logger.warning(f"Received close_short alert but virtual position is not SHORT (size: {state_position_size}) for strategy '{strategy_name}' on account '{acc_name}'. Ignoring.")
                        account_result.update({"success": True, "message": "Current virtual position is not SHORT, ignoring close_short"})
                        results.append(account_result)
                        continue

                    entry_px = float(state_entry_price or 0.0)
                    close_side = "sell" if is_long else "buy"
                    res = client.place_order(
                        product_id=product_id,
                        size=pos_size,
                        side=close_side,
                        order_type="market_order",
                        reduce_only=False
                    )

                    if res.get("success"):
                        order_res = res.get("result", {})
                        try:
                            exit_px = float(order_res.get("average_fill_price") or 0.0)
                        except (ValueError, TypeError):
                            exit_px = 0.0
                        if exit_px <= 0:
                            ticker_data = client.get_ticker(symbol)
                            exit_px = float(ticker_data.get("mark_price") or ticker_data.get("last_price") or 0.0)

                        pnl_val = 0.0
                        if entry_px > 0 and exit_px > 0:
                            direction = 1 if is_long else -1
                            pnl_val = (exit_px - entry_px) * pos_size * contract_value * direction

                        pnl_str = f"PnL: {pnl_val:+.2f} USD"
                        logger.info(f"Successfully closed virtual position for strategy '{strategy_name}' on account '{acc_name}'. Order ID: {order_res.get('id')} - {pnl_str}")
                        
                        # Reset virtual position (short context)
                        if state_id:
                            with app.app_context():
                                state_db = db.session.get(StrategyState, state_id)
                                if state_db:
                                    state_db.position_size = 0.0
                                    state_db.entry_price = None
                                    state_db.current_sl = None
                                    state_db.tp1_price = None
                                    state_db.tp2_price = None
                                    state_db.tp1_hit = False
                                    state_db.tp2_hit = False
                                    db.session.commit()

                        account_result.update({
                            "success": True,
                            "message": f"Closed strategy '{strategy_name}' position of {pos_size} Lots. {pnl_str}",
                            "response": res
                        })
                    else:
                        logger.error(f"Failed to close virtual position for strategy '{strategy_name}' on account '{acc_name}': {res}")
                        account_result.update({"success": False, "message": "Delta API Order Placement Failed", "details": res})

                else:
                    # Account-level default close logic (network calls outside DB context)
                    logger.info(f"Processing close request for {symbol} on account '{acc_name}'...")
                    pos = client.get_position(product_id)
                    if not pos:
                        logger.info(f"No open position found for {symbol} on account '{acc_name}'.")
                        account_result.update({"success": True, "message": "No open position to close"})
                        results.append(account_result)
                        continue

                    try:
                        pos_size = abs(int(float(pos.get("size", 0))))
                    except (ValueError, TypeError):
                        pos_size = 0

                    if pos_size == 0:
                        logger.info(f"Position size is 0 for {symbol} on account '{acc_name}'.")
                        account_result.update({"success": True, "message": "No position size to close"})
                        results.append(account_result)
                        continue

                    pos_side = pos.get("side", "").lower()
                    try:
                        raw_size = float(pos.get("size", 0))
                    except (ValueError, TypeError):
                        raw_size = 0.0

                    is_long = pos_side == "buy" or raw_size > 0
                    is_short = pos_side == "sell" or raw_size < 0

                    # Guard checks to ensure we only close matching position directions
                    if action == "close_long" and not is_long:
                        logger.warning(f"Received close_long alert but position is not LONG (size: {raw_size}) on account '{acc_name}'. Ignoring.")
                        account_result.update({"success": True, "message": "Current position is not LONG, ignoring close_long"})
                        results.append(account_result)
                        continue

                    if action == "close_short" and not is_short:
                        logger.warning(f"Received close_short alert but position is not SHORT (size: {raw_size}) on account '{acc_name}'. Ignoring.")
                        account_result.update({"success": True, "message": "Current position is not SHORT, ignoring close_short"})
                        results.append(account_result)
                        continue

                    # Calculate estimated PnL before closing
                    entry_px = float(pos.get("entry_price") or pos.get("avg_entry_price") or 0)
                    upnl = pos.get("unrealized_pnl") or pos.get("upnl") or pos.get("pnl") or 0.0
                    try:
                        pnl_val = float(upnl)
                    except (ValueError, TypeError):
                        pnl_val = 0.0

                    close_side = "sell" if is_long else "buy"
                    res = client.place_order(
                        product_id=product_id,
                        size=pos_size,
                        side=close_side,
                        order_type="market_order",
                        reduce_only=True
                    )

                    if res.get("success"):
                        order_res = res.get("result", {})
                        try:
                            exit_px = float(order_res.get("average_fill_price") or 0.0)
                        except (ValueError, TypeError):
                            exit_px = 0.0
                            
                        if entry_px > 0 and exit_px > 0:
                            direction = 1 if is_long else -1
                            pnl_val = (exit_px - entry_px) * pos_size * contract_value * direction

                        pnl_str = f"PnL: {pnl_val:+.2f} USD"
                        logger.info(f"Successfully closed position for {symbol} on account '{acc_name}'. Order ID: {order_res.get('id')} - {pnl_str}")
                        
                        # Reset default virtual position (short context)
                        if state_id:
                            with app.app_context():
                                state_db = db.session.get(StrategyState, state_id)
                                if state_db:
                                    state_db.position_size = 0.0
                                    state_db.entry_price = None
                                    state_db.current_sl = None
                                    state_db.tp1_price = None
                                    state_db.tp2_price = None
                                    state_db.tp1_hit = False
                                    state_db.tp2_hit = False
                                    db.session.commit()

                        account_result.update({
                            "success": True,
                            "message": f"Closed position of {pos_size} Lots. {pnl_str}",
                            "response": res
                        })
                    else:
                        logger.error(f"Failed to close position on account '{acc_name}': {res}")
                        account_result.update({"success": False, "message": "Delta API Order Placement Failed", "details": res})
            
            # Check for buy/sell entries
            else:
                if strategy_name:
                    # Strategy-specific Opposing reversal checks
                    if abs(state_position_size) > 0:
                        is_reversal = (action == "buy" and state_position_size < 0) or (action == "sell" and state_position_size > 0)
                        if is_reversal:
                            logger.info(f"Reversal detected! Closing opposing strategy '{strategy_name}' virtual position of size {abs(state_position_size)} first on account '{acc_name}'...")
                            close_side = "buy" if action == "buy" else "sell"
                            close_res = client.place_order(
                                product_id=product_id,
                                size=int(abs(state_position_size)),
                                side=close_side,
                                order_type="market_order",
                                reduce_only=False
                            )
                            if close_res.get("success"):
                                close_order_res = close_res.get("result", {})
                                close_order_id = close_order_res.get("id")
                                try:
                                    exit_px = float(close_order_res.get("average_fill_price") or 0.0)
                                except (ValueError, TypeError):
                                    exit_px = 0.0
                                if exit_px <= 0:
                                    ticker_data = client.get_ticker(symbol)
                                    exit_px = float(ticker_data.get("mark_price") or ticker_data.get("last_price") or 0.0)
                                
                                entry_px = float(state_entry_price or 0.0)
                                is_long = state_position_size > 0
                                pnl_val = 0.0
                                if entry_px > 0 and exit_px > 0:
                                    direction = 1 if is_long else -1
                                    pnl_val = (exit_px - entry_px) * abs(state_position_size) * contract_value * direction
                                
                                pnl_str = f"PnL: {pnl_val:+.2f} USD"
                                logger.info(f"Opposing strategy '{strategy_name}' virtual position closed on account '{acc_name}'. Order ID: {close_order_id} - {pnl_str}")
                                trade_msgs.append(f"Reversal Close Opposing (Order ID: {close_order_id}, {pnl_str})")
                                
                                logger.info(f"Sleeping for 1.5 seconds for margin release...")
                                time.sleep(1.5)
                                # Reset state (short context)
                                if state_id:
                                    with app.app_context():
                                        state_db = db.session.get(StrategyState, state_id)
                                        if state_db:
                                            state_db.position_size = 0.0
                                            state_db.entry_price = None
                                            db.session.commit()
                                state_position_size = 0.0
                            else:
                                logger.error(f"Failed to close opposing virtual position for strategy '{strategy_name}' on account '{acc_name}' during reversal: {close_res}")

                else:
                    # Account-level default Opposing reversal checks (network calls outside DB context)
                    pos = client.get_position(product_id)
                    if pos:
                        try:
                            pos_size = float(pos.get("size", 0))
                        except (ValueError, TypeError):
                            pos_size = 0.0

                        is_reversal = (action == "buy" and pos_size < 0) or (action == "sell" and pos_size > 0)
                        if is_reversal and abs(pos_size) > 0:
                            logger.info(f"Reversal detected! Closing opposing position of size {abs(pos_size)} first on account '{acc_name}'...")
                            close_side = "buy" if action == "buy" else "sell"
                            close_res = client.place_order(
                                product_id=product_id,
                                size=int(abs(pos_size)),
                                side=close_side,
                                order_type="market_order",
                                reduce_only=True
                            )
                            if close_res.get("success"):
                                close_order_res = close_res.get("result", {})
                                close_order_id = close_order_res.get("id")
                                try:
                                    exit_px = float(close_order_res.get("average_fill_price") or 0.0)
                                except (ValueError, TypeError):
                                    exit_px = 0.0
                                if exit_px <= 0:
                                    ticker_data = client.get_ticker(symbol)
                                    exit_px = float(ticker_data.get("mark_price") or ticker_data.get("last_price") or 0.0)
                                
                                entry_px = float(pos.get("entry_price") or pos.get("avg_entry_price") or 0.0)
                                is_long = pos_size > 0
                                pnl_val = 0.0
                                if entry_px > 0 and exit_px > 0:
                                    direction = 1 if is_long else -1
                                    pnl_val = (exit_px - entry_px) * abs(pos_size) * contract_value * direction
                                
                                pnl_str = f"PnL: {pnl_val:+.2f} USD"
                                logger.info(f"Opposing position closed on account '{acc_name}'. Order ID: {close_order_id} - {pnl_str}")
                                trade_msgs.append(f"Reversal Close Opposing (Order ID: {close_order_id}, {pnl_str})")
                                
                                logger.info(f"Sleeping for 1.5 seconds for margin release...")
                                time.sleep(1.5)
                                if state_id:
                                    with app.app_context():
                                        state_db = db.session.get(StrategyState, state_id)
                                        if state_db:
                                            state_db.position_size = 0.0
                                            state_db.entry_price = None
                                            db.session.commit()
                            else:
                                logger.error(f"Failed to close opposing position on account '{acc_name}' during reversal: {close_res}")

                # Fetch ticker details for price (network call, outside DB context)
                ticker_data = client.get_ticker(symbol)
                price_str = ticker_data.get("mark_price") or ticker_data.get("last_price") if ticker_data else None
                try:
                    price = float(price_str)
                except (ValueError, TypeError):
                    logger.error(f"Invalid price value received for {symbol} on account '{acc_name}': '{price_str}'")
                    account_result.update({"success": False, "message": "Failed to fetch symbol price"})
                    results.append(account_result)
                    continue

                # Fetch account available balance (network call, outside DB context)
                balance, asset = client.get_available_balance()
                if balance <= 0:
                    logger.warning(f"Available balance is 0 on account '{acc_name}'.")
                    account_result.update({"success": False, "message": "Available balance is 0"})
                    results.append(account_result)
                    continue

                # Calculate quantity and lots
                qty_lots = None
                sizing_desc = ""
                lot_value_usd = price * contract_value
                
                if lot_value_usd <= 0:
                    logger.error(f"Invalid lot value calculation for account '{acc_name}'")
                    account_result.update({"success": False, "message": "Invalid lot value calculation"})
                    results.append(account_result)
                    continue
                    
                if payload and ("quantity" in payload or "qty" in payload):
                    payload_qty = payload.get("quantity") or payload.get("qty")
                    if payload_qty is not None:
                        try:
                            qty_base = float(payload_qty)
                            qty_lots = int(math.floor(qty_base / contract_value))
                            sizing_desc = f"Quantity from payload = {qty_base} (Lots = {qty_lots})"
                        except (ValueError, TypeError) as e:
                            logger.error(f"Invalid quantity parameter in payload: {payload_qty}. Falling back to account sizing.")
                            
                if qty_lots is None:
                    if strategy_name:
                        # Strategy Sizing Mode
                        leverage = strategy_leverage
                        sizing_type = strategy_sizing_type
                        fixed_amount = strategy_fixed_amount
                        balance_buffer_pct = strategy_balance_buffer_pct
                        
                        if sizing_type == "fixed":
                            if fixed_amount > balance:
                                logger.warning(f"Insufficient balance on account '{acc_name}' for strategy '{strategy_name}': Fixed Margin of {fixed_amount} {asset} exceeds balance {balance} {asset}.")
                                account_result.update({"success": False, "message": f"Insufficient balance (need {fixed_amount} {asset}, have {balance} {asset})"})
                                results.append(account_result)
                                continue
                            buying_power = fixed_amount * leverage
                            sizing_desc = f"Strategy '{strategy_name}' Fixed Margin = {fixed_amount} {asset}"
                        else:
                            buying_power = balance * leverage * (balance_buffer_pct / 100.0)
                            sizing_desc = f"Strategy '{strategy_name}' Buffer = {balance_buffer_pct}%"
                    else:
                        # Account-level Default Sizing Mode
                        leverage = acc["leverage"]
                        sizing_type = acc.get("sizing_type") or "percentage"
                        fixed_amount_val = acc.get("fixed_amount")
                        fixed_amount = float(fixed_amount_val) if fixed_amount_val is not None else 10.0
                        
                        if sizing_type == "fixed":
                            if fixed_amount > balance:
                                logger.warning(f"Insufficient balance on account '{acc_name}': Fixed Margin of {fixed_amount} {asset} exceeds balance {balance} {asset}.")
                                account_result.update({"success": False, "message": f"Insufficient balance (need {fixed_amount} {asset}, have {balance} {asset})"})
                                results.append(account_result)
                                continue
                            buying_power = fixed_amount * leverage
                            sizing_desc = f"Fixed Margin = {fixed_amount} {asset}"
                        else:
                            buffer_pct = float(acc.get("balance_buffer_pct", 55.0))
                            buying_power = balance * leverage * (buffer_pct / 100.0)
                            sizing_desc = f"Buffer = {buffer_pct}%"
                            
                    qty_lots = int(math.floor(buying_power / lot_value_usd))
                    
                # Required margin check
                used_leverage = strategy_leverage if strategy_name else acc["leverage"]
                required_margin = (qty_lots * lot_value_usd) / used_leverage
                if required_margin > balance:
                    logger.warning(f"Insufficient balance on account '{acc_name}': Required margin of {required_margin:.4f} {asset} for {qty_lots} lots exceeds available balance {balance} {asset}.")
                    account_result.update({"success": False, "message": f"Insufficient balance (need {required_margin:.2f} {asset} margin, have {balance:.2f} {asset})"})
                    results.append(account_result)
                    continue
                    
                logger.info(f"Sizing details for account '{acc_name}': Balance = {balance} {asset}, Leverage = {used_leverage}x, "
                            f"{sizing_desc}, Lot USD Value = {lot_value_usd:.4f}, Calculated Qty = {qty_lots} Lots")

                if qty_lots <= 0:
                    logger.warning(f"Insufficient balance ({balance} {asset}) for leverage {used_leverage}x to open even 1 lot on account '{acc_name}'.")
                    account_result.update({"success": False, "message": "Insufficient balance for 1 lot"})
                    results.append(account_result)
                    continue

                # Execute market order to enter trade (network call, outside DB context)
                res = client.place_order(
                    product_id=product_id,
                    size=qty_lots,
                    side=action,
                    order_type="market_order",
                    reduce_only=False
                )

                if res.get("success"):
                    order_res = res.get("result", {})
                    try:
                        fill_px = float(order_res.get("average_fill_price") or order_res.get("price") or 0.0)
                    except (ValueError, TypeError):
                        fill_px = 0.0
                    if fill_px <= 0:
                        fill_px = price
                        
                    # Update virtual position tracking in database (short context)
                    if state_id:
                        with app.app_context():
                            state_db = db.session.get(StrategyState, state_id)
                            if state_db:
                                signed_size = qty_lots if action == "buy" else -qty_lots
                                state_db.position_size = signed_size
                                state_db.entry_price = fill_px
                                state_db.current_sl = None
                                state_db.tp1_price = None
                                state_db.tp2_price = None
                                state_db.tp1_hit = False
                                state_db.tp2_hit = False
                                db.session.commit()
                                logger.info(f"Updated StrategyState: {symbol} (Strategy ID: {strategy_id}, Size: {signed_size}, Entry: {fill_px})")
                        
                    order_id = order_res.get('id')
                    logger.info(f"Successfully entered position for {symbol} on account '{acc_name}'. Order ID: {order_id}")
                    trade_msgs.append(f"Entered {qty_lots} Lots @ {fill_px:.2f} (Order ID: {order_id})")
                    account_result.update({
                        "success": True,
                        "message": " | ".join(trade_msgs),
                        "response": res
                    })
                else:
                    logger.error(f"Failed to place order on account '{acc_name}': {res}")
                    fail_msg = "Delta API Order Placement Failed"
                    if trade_msgs:
                        fail_msg = " | ".join(trade_msgs) + f" | Entry Failed: {res.get('error') or 'unknown'}"
                    account_result.update({"success": False, "message": fail_msg, "details": res})

        except Exception as acc_e:
            logger.exception(f"Exception processing webhook in background for account '{acc_name}': {acc_e}")
            account_result.update({"success": False, "message": "Internal processing exception", "error": str(acc_e)})
            
        results.append(account_result)
        
    # Logging results to database (short context)
    status = "success"
    details_list = []
    success_count = sum(1 for r in results if r["success"])
    
    if len(results) == 0:
        status = "failed"
        details_list.append("No active accounts to execute.")
    elif success_count == len(results):
        status = "success"
    elif success_count == 0:
        status = "failed"
    else:
        status = "partial"
        
    for r in results:
        name = r["name"]
        if r["success"]:
            msg = r.get("message") or "Order placed successfully"
            order_id = r.get("response", {}).get("result", {}).get("id")
            if order_id and f"Order ID: {order_id}" not in msg:
                details_list.append(f"{name}: Success (Order ID: {order_id}) - {msg}")
            else:
                details_list.append(f"{name}: Success ({msg})")
        else:
            msg = r.get("message") or r.get("error") or "Unknown error"
            details_list.append(f"{name}: Failed ({msg})")
            
    details_str = "\n".join(details_list)
    with app.app_context():
        log_entry = TradeLog(
            ticker=ticker,
            action=action,
            source=source,
            status=status,
            details=details_str
        )
        db.session.add(log_entry)
        db.session.commit()
    logger.info(f"Saved TradeLog entry to database. Status: {status}")
    
    # Send Notification
    title_emoji = "🟢" if status == "success" else ("🟡" if status == "partial" else "🔴")
    title = f"{title_emoji} Trade Alert: {ticker} ({action.upper()})"
    notification_message = (
        f"<b>Source:</b> {source}\n"
        f"<b>Status:</b> {status.upper()}\n\n"
        f"<b>Details:</b>\n<pre>{details_str}</pre>"
    )
    status_color = 1096065 if status == "success" else (16498468 if status == "partial" else 15680580)
    send_notification(title, notification_message, status_color, symbol=ticker)
    
    # Trigger consecutive loss checks for each active account
    try:
        from loss_analyzer import check_and_analyze_consecutive_losses
        with app.app_context():
            active_accounts = Account.query.filter_by(is_active=True).all()
            for acc in active_accounts:
                check_and_analyze_consecutive_losses(app, acc)
    except Exception as la_err:
        logger.error(f"Error running consecutive loss checks: {la_err}")
        
    logger.info("Background trade execution completed.")
    return results

@app.route("/webhook", methods=["POST"])
def webhook():
    """TradingView webhook endpoint executing trades on all active accounts."""
    try:
        payload = request.get_json(silent=True)
        if not payload:
            logger.warning("Received request with missing or invalid JSON body.")
            log_entry = TradeLog(
                ticker="UNKNOWN",
                action="UNKNOWN",
                source="webhook",
                status="failed",
                details="Received request with missing or invalid JSON body."
            )
            db.session.add(log_entry)
            db.session.commit()
            return jsonify({"status": "error", "message": "Missing JSON body"}), 400

        logger.info(f"Incoming webhook payload: {payload}")

        # 1. Validate passphrase from database
        passphrase = request.args.get("passphrase") or payload.get("passphrase")
        passphrase_setting = GlobalSetting.query.filter_by(key="passphrase").first()
        db_passphrase = passphrase_setting.value if passphrase_setting else Config.PASSPHRASE
        
        if passphrase != db_passphrase:
            logger.warning(f"Unauthorized access attempt with invalid passphrase: '{passphrase}'")
            log_entry = TradeLog(
                ticker=payload.get("ticker", "UNKNOWN"),
                action=payload.get("action", "UNKNOWN"),
                source="webhook",
                status="failed",
                details=f"Unauthorized access attempt with invalid passphrase: '{passphrase}'"
            )
            db.session.add(log_entry)
            db.session.commit()
            return jsonify({"status": "error", "message": "Unauthorized"}), 401

        # 2. Extract symbol and action
        ticker = payload.get("ticker")
        action = payload.get("action", "").lower()

        if not ticker:
            logger.error("Missing 'ticker' in webhook payload.")
            log_entry = TradeLog(
                ticker="UNKNOWN",
                action=action or "UNKNOWN",
                source="webhook",
                status="failed",
                details="Missing 'ticker' in webhook payload."
            )
            db.session.add(log_entry)
            db.session.commit()
            return jsonify({"status": "error", "message": "Missing 'ticker'"}), 400

        if action not in ["buy", "sell", "close_long", "close_short"]:
            logger.error(f"Invalid 'action' in payload: '{action}'")
            log_entry = TradeLog(
                ticker=ticker,
                action=action or "UNKNOWN",
                source="webhook",
                status="failed",
                details=f"Invalid 'action' in payload: '{action}'. Must be buy, sell, close_long, or close_short"
            )
            db.session.add(log_entry)
            db.session.commit()
            return jsonify({"status": "error", "message": "Invalid 'action'. Must be buy, sell, close_long, or close_short"}), 400

        # 3. Fetch active accounts (optionally filtered by target account in payload/query)
        target_account_name = request.args.get("account") or request.args.get("account_name") or payload.get("account") or payload.get("account_name")
        target_account_id = request.args.get("account_id") or payload.get("account_id")
        
        # Extract strategy name from query parameter or payload
        target_strategy_name = request.args.get("strategy") or request.args.get("strategy_name") or payload.get("strategy") or payload.get("strategy_name")
        if target_strategy_name:
            payload["strategy"] = target_strategy_name
            
        if target_account_id is not None:
            try:
                acc_id = int(target_account_id)
                if acc_id == 0:
                    active_accounts = []
                else:
                    active_accounts = Account.query.filter_by(id=acc_id, is_active=True).all()
            except ValueError:
                active_accounts = []
        elif target_account_name:
            active_accounts = Account.query.filter(
                Account.name.ilike(target_account_name),
                Account.is_active == True
            ).all()
        else:
            active_accounts = Account.query.filter_by(is_active=True).all()

        if not active_accounts:
            should_fallback = False
            if target_account_id is not None:
                try:
                    should_fallback = (int(target_account_id) == 0)
                except ValueError:
                    pass
            elif target_account_name:
                should_fallback = (target_account_name.lower() in ["environment default", "environment_default", "default"])
            else:
                should_fallback = True

            if should_fallback and Config.API_KEY and Config.API_SECRET:
                logger.info("No active accounts configured in database. Falling back to environment API credentials.")
                fallback_account = Account(
                    id=0,
                    name="Environment Default",
                    api_key=Config.API_KEY,
                    api_secret=Config.API_SECRET,
                    leverage=Config.DEFAULT_LEVERAGE,
                    balance_buffer_pct=Config.BALANCE_BUFFER_PCT * 100.0,
                    sizing_type="percentage",
                    fixed_amount=10.0,
                    is_active=True
                )
                active_accounts = [fallback_account]
            else:
                filter_desc = f" (Filter: Name={target_account_name}, ID={target_account_id})" if (target_account_name or target_account_id) else ""
                logger.warning(f"No active accounts found in database matching criteria{filter_desc}. Skipping webhook execution.")
                log_entry = TradeLog(
                    ticker=ticker,
                    action=action,
                    source="webhook",
                    status="ignored",
                    details=f"No active accounts found matching criteria{filter_desc}."
                )
                db.session.add(log_entry)
                db.session.commit()
                return jsonify({"status": "success", "message": f"No active accounts matched{filter_desc}"}), 200

        # 4. Extract account data into plain dictionaries to pass to background thread
        accounts_data = []
        for account in active_accounts:
            accounts_data.append({
                "id": account.id,
                "name": account.name,
                "api_key": account.api_key,
                "api_secret": account.api_secret,
                "leverage": account.leverage,
                "balance_buffer_pct": account.balance_buffer_pct,
                "sizing_type": account.sizing_type,
                "fixed_amount": account.fixed_amount
            })

        # 5. Launch background thread for trade execution
        if os.getenv("FLASK_ENV") == "testing":
            # Run synchronously in testing to keep assertions deterministic
            results = execute_trades_background(accounts_data, ticker, action, "webhook", payload)
            return jsonify({"status": "success", "results": results}), 200
        else:
            import threading
            thread = threading.Thread(
                target=execute_trades_background,
                args=(accounts_data, ticker, action, "webhook"),
                kwargs={"payload": payload}
            )
            thread.start()

            return jsonify({
                "status": "success",
                "message": "Webhook signal received. Execution started in background.",
                "accounts_count": len(accounts_data)
            }), 200

    except Exception as e:
        logger.exception(f"Unhandled exception in webhook execution: {e}")
        try:
            log_entry = TradeLog(
                ticker=payload.get("ticker", "UNKNOWN") if payload else "UNKNOWN",
                action=payload.get("action", "UNKNOWN") if payload else "UNKNOWN",
                source="webhook",
                status="failed",
                details=f"Unhandled exception: {e}"
            )
            db.session.add(log_entry)
            db.session.commit()
        except Exception as db_e:
            logger.error(f"Failed to save error log to DB: {db_e}")
        return jsonify({"status": "error", "message": "Internal Server Error", "exception": str(e)}), 500

def run_migrations():
    from sqlalchemy import text
    try:
        # 1. Add sizing_type
        try:
            db.session.execute(text("ALTER TABLE accounts ADD COLUMN sizing_type VARCHAR(20) DEFAULT 'percentage' NOT NULL"))
            db.session.commit()
            logger.info("Database migration: added sizing_type column to accounts table.")
        except Exception as e:
            db.session.rollback()
            logger.debug(f"sizing_type migration status: {e}")
            
        # 2. Add fixed_amount
        try:
            db.session.execute(text("ALTER TABLE accounts ADD COLUMN fixed_amount FLOAT DEFAULT 10.0 NOT NULL"))
            db.session.commit()
            logger.info("Database migration: added fixed_amount column to accounts table.")
        except Exception as e:
            db.session.rollback()
            logger.debug(f"fixed_amount migration status: {e}")
            
        # 3. Add daily_loss_limit
        try:
            db.session.execute(text("ALTER TABLE accounts ADD COLUMN daily_loss_limit FLOAT NULL"))
            db.session.commit()
            logger.info("Database migration: added daily_loss_limit column to accounts table.")
        except Exception as e:
            db.session.rollback()
            logger.debug(f"daily_loss_limit migration status: {e}")
            
        # 4. Add is_circuit_broken
        try:
            db.session.execute(text("ALTER TABLE accounts ADD COLUMN is_circuit_broken BOOLEAN DEFAULT FALSE NOT NULL"))
            db.session.commit()
            logger.info("Database migration: added is_circuit_broken column to accounts table.")
        except Exception as e:
            db.session.rollback()
            logger.debug(f"is_circuit_broken migration status: {e}")
            
        # 5. Add local_strategy_enabled
        try:
            db.session.execute(text("ALTER TABLE accounts ADD COLUMN local_strategy_enabled BOOLEAN DEFAULT FALSE NOT NULL"))
            db.session.commit()
            logger.info("Database migration: added local_strategy_enabled column to accounts table.")
        except Exception as e:
            db.session.rollback()
            logger.debug(f"local_strategy_enabled migration status: {e}")

        # 6. Create strategies table and add columns
        try:
            # Check if strategies table exists
            db.session.execute(text("SELECT 1 FROM strategies LIMIT 1"))
        except Exception:
            db.session.rollback()
            try:
                db.session.execute(text("""
                    CREATE TABLE strategies (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        account_id INTEGER NOT NULL REFERENCES accounts(id) ON DELETE CASCADE,
                        name VARCHAR(100) NOT NULL,
                        is_active BOOLEAN DEFAULT TRUE NOT NULL,
                        sizing_type VARCHAR(20) DEFAULT 'percentage' NOT NULL,
                        balance_buffer_pct FLOAT DEFAULT 10.0 NOT NULL,
                        fixed_amount FLOAT DEFAULT 10.0 NOT NULL,
                        leverage INTEGER DEFAULT 50 NOT NULL,
                        UNIQUE(account_id, name)
                    )
                """))
                db.session.commit()
                logger.info("Database migration: created strategies table.")
            except Exception as table_e:
                db.session.rollback()
                logger.error(f"Failed to create strategies table: {table_e}")

        # 7. Add strategy_id to strategy_states
        try:
            db.session.execute(text("ALTER TABLE strategy_states ADD COLUMN strategy_id INTEGER REFERENCES strategies(id) ON DELETE CASCADE"))
            db.session.commit()
            logger.info("Database migration: added strategy_id column to strategy_states table.")
        except Exception as e:
            db.session.rollback()
            logger.debug(f"strategy_id migration status: {e}")

        # 8. Add manual_exit_detected to strategy_states
        try:
            db.session.execute(text("ALTER TABLE strategy_states ADD COLUMN manual_exit_detected INTEGER DEFAULT 0 NOT NULL"))
            db.session.commit()
            logger.info("Database migration: added manual_exit_detected column to strategy_states table.")
        except Exception as e:
            db.session.rollback()
            logger.debug(f"manual_exit_detected migration status: {e}")
            
        # 9. Create index on strategy_states for fast queries
        try:
            db.session.execute(text("CREATE INDEX IF NOT EXISTS idx_strategy_states_lookup ON strategy_states (account_id, symbol, strategy_id)"))
            db.session.commit()
            logger.info("Database migration: created index idx_strategy_states_lookup on strategy_states.")
        except Exception as e:
            db.session.rollback()
            logger.debug(f"index creation status: {e}")
    except Exception as e:
        logger.error(f"Migration failed: {e}")

# Initialize database tables on startup (unless running tests)
if os.getenv("FLASK_ENV") != "testing":
    with app.app_context():
        db.create_all()
        # Run database migrations for any new columns
        run_migrations()
        # Initialize default passphrase in database if not present
        passphrase_setting = GlobalSetting.query.filter_by(key="passphrase").first()
        if not passphrase_setting:
            initial_passphrase = Config.PASSPHRASE or "my_secure_passphrase"
            db.session.add(GlobalSetting(key="passphrase", value=initial_passphrase))
            db.session.commit()
            logger.info(f"Initialized default passphrase in database: {initial_passphrase}")
        
        # Initialize default local_bot_dry_run setting to true if not present
        dry_run_setting = GlobalSetting.query.filter_by(key="local_bot_dry_run").first()
        if not dry_run_setting:
            db.session.add(GlobalSetting(key="local_bot_dry_run", value="true"))
            db.session.commit()
            logger.info("Initialized default local_bot_dry_run setting to true.")
        
        # Dispose the engine pool so that Gunicorn workers do not inherit open connection sockets
        try:
            db_uri = app.config.get("SQLALCHEMY_DATABASE_URI") or ""
            if "sqlite" not in db_uri:
                db.engine.dispose()
        except Exception as e:
            logger.warning(f"Failed to dispose engine pool at startup: {e}")

if __name__ == "__main__":
    port = int(os.getenv("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
