import time
import logging
import math
import os
import pandas as pd
import numpy as np
from datetime import datetime
from models import db, Account, StrategyState, TradeLog
from delta_client import DeltaClient
from strategy_logic import evaluate_strategy

logger = logging.getLogger(__name__)

# Constants matching Strategy optimized parameters
SYMBOL = "ETHUSD.P"
RESOLUTION = "5"  # 5 minutes
TP1_RR = 1.5
TP2_RR = 3.0
USE_BE = True
USE_LIQ_EXIT = True
USE_ZLSMA_EXIT = True

def send_strategy_notification(app, title, message, status_color=5763719, symbol=None):
    """Dispatches strategy alerts to Discord/Telegram using existing notification helper in app."""
    try:
        from app import send_notification
        from models import GlobalSetting
        
        # Prepend Dry-Run tag if enabled
        dry_run = True
        if os.getenv("FLASK_ENV") != "testing":
            try:
                with app.app_context():
                    dry_run_s = GlobalSetting.query.filter_by(key="local_bot_dry_run").first()
                    if dry_run_s:
                        dry_run = dry_run_s.value.lower() == "true"
            except Exception as e:
                logger.error(f"Failed to read local_bot_dry_run: {e}")
        else:
            dry_run = False
            
        if dry_run:
            title = f"🔍 [Dry-Run] {title}"
            
        if os.getenv("FLASK_ENV") == "testing":
            send_notification(title, message, status_color, symbol=symbol)
        else:
            with app.app_context():
                send_notification(title, message, status_color, symbol=symbol)
    except Exception as e:
        logger.error(f"Failed to send strategy notification: {e}")

def send_entry_email_alert(app, account_name, symbol, side, qty_lots, price, stop_px, sl_dist, is_reentry=False):
    """Sends an email alert to the user when a new entry (or re-entry) is detected."""
    reentry_tag = " [RE-ENTRY]" if is_reentry else ""
    subject = f"🔔 Local Strategy: {side.upper()} Entry{reentry_tag} for {symbol}"
    
    # Use strategy runner parameters TP1_RR and TP2_RR
    tp1_price = price + sl_dist * TP1_RR if side.lower() == "buy" else price - sl_dist * TP1_RR
    tp2_price = price + sl_dist * TP2_RR if side.lower() == "buy" else price - sl_dist * TP2_RR
    
    html_body = f"""
    <h3>New Local Strategy Signal Detected</h3>
    <p><b>Account:</b> {account_name}</p>
    <p><b>Symbol:</b> {symbol}</p>
    <p><b>Action:</b> {side.upper()}{reentry_tag}</p>
    <p><b>Signal Price:</b> {price:.2f}</p>
    <p><b>Suggested Quantity:</b> {qty_lots} contracts</p>
    <p><b>Stop Loss:</b> {stop_px:.2f} (Dist: {sl_dist:.2f})</p>
    <p><b>Take Profit 1:</b> {tp1_price:.2f}</p>
    <p><b>Take Profit 2:</b> {tp2_price:.2f}</p>
    <br>
    <p><i>Note: The bot is configured to NOT enter trades automatically. Please execute this trade manually on Delta Exchange.</i></p>
    """
    try:
        from app import send_email_alert
        if os.getenv("FLASK_ENV") == "testing":
            logger.info(f"Simulating email alert in testing: {subject}")
            return True
        else:
            with app.app_context():
                return send_email_alert(subject, html_body)
    except Exception as e:
        logger.error(f"Failed to send email alert: {e}")
        return False

def update_strategy_state(app, state_id, updates):
    """Helper to update StrategyState inside a short-lived db context."""
    if os.getenv("FLASK_ENV") == "testing":
        state = db.session.get(StrategyState, state_id)
        if state:
            for k, v in updates.items():
                setattr(state, k, v)
            db.session.commit()
    else:
        with app.app_context():
            state = db.session.get(StrategyState, state_id)
            if state:
                for k, v in updates.items():
                    setattr(state, k, v)
                db.session.commit()

def log_trade(app, ticker, action, source, status, details):
    """Helper to add TradeLog entry inside a short-lived db context."""
    if os.getenv("FLASK_ENV") == "testing":
        trade_log = TradeLog(
            ticker=ticker,
            action=action,
            source=source,
            status=status,
            details=details
        )
        db.session.add(trade_log)
        db.session.commit()
    else:
        with app.app_context():
            trade_log = TradeLog(
                ticker=ticker,
                action=action,
                source=source,
                status=status,
                details=details
            )
            db.session.add(trade_log)
            db.session.commit()

def append_signal_to_excel(account_name, sig_type, price, quantity, stop_loss=None, tp1=None, tp2=None):
    """Appends dry-run signal details to a local Excel file 'local_bot_signals.xlsx'."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import os
    from datetime import datetime
    
    project_dir = os.path.dirname(os.path.abspath(__file__))
    filepath = os.path.join(project_dir, "local_bot_signals.xlsx")
    
    header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    
    headers = [
        "TIMESTAMP (UTC)", "ACCOUNT NAME", "SIGNAL TYPE", "PRICE (USD)", 
        "QUANTITY (LOTS)", "STOP LOSS (USD)", "TAKE PROFIT 1 (USD)", "TAKE PROFIT 2 (USD)"
    ]
    
    if not os.path.exists(filepath):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Local Bot Signals"
        ws.views.sheetView[0].showGridLines = True
        
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        ws.row_dimensions[1].height = 28
    else:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
    timestamp_str = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
    row_data = [
        timestamp_str,
        account_name,
        sig_type,
        price,
        quantity,
        stop_loss if stop_loss else "",
        tp1 if tp1 else "",
        tp2 if tp2 else ""
    ]
    
    next_row = ws.max_row + 1
    for col_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=next_row, column=col_idx, value=val)
        cell.font = data_font
        cell.border = thin_border
        
        if col_idx in [1, 2]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 3:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if val == "BUY":
                cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                cell.font = Font(name="Segoe UI", size=10, bold=True, color="065F46")
            elif val == "SELL":
                cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                cell.font = Font(name="Segoe UI", size=10, bold=True, color="991B1B")
            elif val == "SL":
                cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                cell.font = Font(name="Segoe UI", size=10, bold=True, color="92400E")
            else:
                cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
                cell.font = Font(name="Segoe UI", size=10, bold=True, color="1E40AF")
        else:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            if val != "":
                cell.number_format = "$#,##0.00" if col_idx != 5 else "#,##0"
                
        ws.row_dimensions[next_row].height = 20
        
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    wb.save(filepath)

def run_strategy_for_account(app, account, client, symbol=SYMBOL, strategy_type="chandelier_exit", ce_length=22, ce_mult=3.0, zlsma_length=32, vol_length=20, vol_mult=1.15, tp1_rr=1.5, tp2_rr=3.0, use_be=True, use_liq=True, use_zlsma=True, resolution="5", zl_length=70, zl_mult=1.2, ai_speed=14, ai_atr_len=14, ai_atr_mult=2.0):
    # Support both dictionary and db object
    if isinstance(account, dict):
        account_id = account["id"]
        account_name = account["name"]
        account_leverage = account["leverage"]
        is_circuit_broken = account["is_circuit_broken"]
        balance_buffer_pct = float(account.get("balance_buffer_pct", 55.0))
        sizing_type = account.get("sizing_type", "percentage")
        fixed_amount = float(account.get("fixed_amount", 10.0))
    else:
        account_id = account.id
        account_name = account.name
        account_leverage = account.leverage
        is_circuit_broken = account.is_circuit_broken
        balance_buffer_pct = float(account.balance_buffer_pct) if account.balance_buffer_pct is not None else 55.0
        sizing_type = account.sizing_type
        fixed_amount = float(account.fixed_amount) if account.fixed_amount is not None else 10.0

    # 1. Fetch product details
    product = client.get_product_by_symbol(symbol)
    if not product:
        logger.error(f"[{account_name}] Symbol '{symbol}' not found on Delta. Skipping.")
        return
        
    product_id = product.get("id")
    contract_val = float(product.get("contract_value", 0.01))
    tick_size = float(product.get("tick_size", 0.05))
    
    # 2. Get or create StrategyState record inside short-lived context
    if os.getenv("FLASK_ENV") == "testing":
        state_db = StrategyState.query.filter_by(account_id=account_id, symbol=symbol, strategy_id=None).first()
        if not state_db:
            state_db = StrategyState(account_id=account_id, symbol=symbol, strategy_id=None, position_size=0.0)
            db.session.add(state_db)
            db.session.commit()
            
        state_id = state_db.id
        position_size = state_db.position_size
        entry_price = state_db.entry_price
        sl_dist = state_db.sl_dist
        tp1_price = state_db.tp1_price
        tp2_price = state_db.tp2_price
        tp1_hit = state_db.tp1_hit
        tp2_hit = state_db.tp2_hit
        current_sl = state_db.current_sl
        last_signal_time = state_db.last_signal_time
        manual_exit_detected = getattr(state_db, "manual_exit_detected", 0)
    else:
        with app.app_context():
            state_db = StrategyState.query.filter_by(account_id=account_id, symbol=symbol, strategy_id=None).first()
            if not state_db:
                state_db = StrategyState(account_id=account_id, symbol=symbol, strategy_id=None, position_size=0.0)
                db.session.add(state_db)
                db.session.commit()
                
            state_id = state_db.id
            position_size = state_db.position_size
            entry_price = state_db.entry_price
            sl_dist = state_db.sl_dist
            tp1_price = state_db.tp1_price
            tp2_price = state_db.tp2_price
            tp1_hit = state_db.tp1_hit
            tp2_hit = state_db.tp2_hit
            current_sl = state_db.current_sl
            last_signal_time = state_db.last_signal_time
            manual_exit_detected = getattr(state_db, "manual_exit_detected", 0)
        
    to_time = int(time.time())
    
    # Safe resolution parsing (supports 1D, 1W, 1M resolutions)
    try:
        res_mins = int(resolution)
    except ValueError:
        res_upper = str(resolution).upper()
        if res_upper == "1D":
            res_mins = 1440
        elif res_upper == "1W":
            res_mins = 10080
        elif res_upper == "1M":
            res_mins = 43200
        else:
            res_mins = 5
            
    # Fetch 2,000 candles of history for 100% mathematical indicators warmup
    from_time = to_time - (2000 * res_mins * 60)
    exchange_symbol = product.get("symbol", symbol)
    query_str = f"symbol={exchange_symbol}&resolution={resolution}&from={from_time}&to={to_time}"
    
    response = client._request("GET", "/v2/chart/history", query_string=query_str, is_private=False)
    if not response.get("success"):
        logger.error(f"[{account_name}] Failed to fetch candles: {response.get('error')}")
        return
        
    result = response.get("result", {})
    close_prices = result.get("c", [])
    if len(close_prices) < 100:
        logger.warning(f"[{account_name}] Insufficient candles returned: {len(close_prices)}. Need at least 100.")
        return
        
    df = pd.DataFrame({
        "close": [float(x) for x in result.get("c", [])],
        "high": [float(x) for x in result.get("h", [])],
        "low": [float(x) for x in result.get("l", [])],
        "open": [float(x) for x in result.get("o", [])],
        "volume": [float(x) for x in result.get("v", [])],
        "time": [int(x) for x in result.get("t", [])]
    })
    
    # Calculate indicators
    signals = evaluate_strategy(
        df, strategy_type=strategy_type, 
        ce_length=ce_length, ce_mult=ce_mult, 
        zlsma_length=zlsma_length, vol_length=vol_length, vol_mult=vol_mult,
        zl_length=zl_length, zl_mult=zl_mult,
        ai_speed=ai_speed, ai_atr_len=ai_atr_len, ai_atr_mult=ai_atr_mult
    )
    
    # Determine which candle index to evaluate
    if os.getenv("FLASK_ENV") == "testing":
        candle_idx = len(df) - 2
    else:
        current_time_epoch = int(time.time())
        resolution_seconds = res_mins * 60
        completed_df = df[df["time"] + resolution_seconds <= current_time_epoch]
        if completed_df.empty:
            logger.warning(f"[{account_name}] No completed candles found in data.")
            return
        candle_idx = completed_df.index[-1]
        
    last_completed_time = df["time"].iloc[candle_idx]
    last_close = df["close"].iloc[candle_idx]
    
    # Fetch global dry-run setting
    dry_run = True
    if os.getenv("FLASK_ENV") != "testing":
        try:
            with app.app_context():
                from models import GlobalSetting
                dry_run_s = GlobalSetting.query.filter_by(key="local_bot_dry_run").first()
                if dry_run_s:
                    dry_run = dry_run_s.value.lower() == "true"
        except Exception as e:
            logger.error(f"Failed to read local_bot_dry_run: {e}")
    else:
        dry_run = False
 
    if dry_run:
        # Mock client.get_position to return our local virtual position size
        def mock_get_position(prod_id):
            return {
                "size": position_size,
                "entry_price": entry_price or 0.0,
                "product_id": prod_id
            }
        client.get_position = mock_get_position
 
        # Wrap client.place_order to prevent real execution and log to DB instead
        def mock_place_order(prod_id, size, side, order_type="market_order", reduce_only=False):
            logger.info(f"[Dry-Run Mode] Simulating local strategy order for account '{account_name}': {side.upper()} {size} lots on product ID {prod_id} (Reduce Only: {reduce_only})")
            
            # Fetch current ticker mark price
            ticker_data_live = client.get_ticker(symbol)
            current_price_live = float(ticker_data_live.get("mark_price") or ticker_data_live.get("last_price") or last_close)
            
            # Determine signal type
            if not reduce_only:
                sig_type = "BUY" if side == "buy" else "SELL"
            else:
                if position_size > 0:
                    if side == "sell":
                        if abs(size - int(math.floor(position_size * 0.5))) <= 1:
                            sig_type = "TP1"
                        elif abs(size - int(math.floor(position_size * 0.3))) <= 1:
                            sig_type = "TP2"
                        elif current_price_live <= (current_sl or 0.0):
                            sig_type = "SL"
                        else:
                            sig_type = "EXIT_ZLSMA" if use_zlsma else "EXIT_LIQUIDITY"
                    else:
                        sig_type = "BUY"
                elif position_size < 0:
                    if side == "buy":
                        if abs(size - int(math.floor(abs(position_size) * 0.5))) <= 1:
                            sig_type = "TP1"
                        elif abs(size - int(math.floor(abs(position_size) * 0.3))) <= 1:
                            sig_type = "TP2"
                        elif current_price_live >= (current_sl or 0.0):
                            sig_type = "SL"
                        else:
                            sig_type = "EXIT_ZLSMA" if use_zlsma else "EXIT_LIQUIDITY"
                    else:
                        sig_type = "SELL"
                else:
                    sig_type = "CLOSE_EMERGENCY"
            
            # Save simulated signal log to database (short context)
            with app.app_context():
                try:
                    from models import LocalSignalLog
                    state_db = StrategyState.query.filter_by(account_id=account_id, symbol=symbol, strategy_id=None).first()
                    
                    stop_loss_val = None
                    tp1_val = None
                    tp2_val = None
                    if state_db:
                        stop_loss_val = state_db.current_sl
                        tp1_val = state_db.tp1_price
                        tp2_val = state_db.tp2_price
                        
                    if sig_type in ["BUY", "SELL"]:
                        stop_px = signals["long_stop"][candle_idx] if sig_type == "BUY" else signals["short_stop"][candle_idx]
                        if not np.isnan(stop_px):
                            stop_loss_val = stop_px
                            sl_dist_calc = max(abs(current_price_live - stop_px), tick_size)
                            tp1_val = current_price_live + sl_dist_calc * tp1_rr if sig_type == "BUY" else current_price_live - sl_dist_calc * tp1_rr
                            tp2_val = current_price_live + sl_dist_calc * tp2_rr if sig_type == "BUY" else current_price_live - sl_dist_calc * tp2_rr
                    
                    log = LocalSignalLog(
                        account_id=account_id,
                        account_name=account_name,
                        signal_type=sig_type,
                        price=current_price_live,
                        quantity=float(size),
                        stop_loss=float(stop_loss_val) if stop_loss_val else None,
                        take_profit_1=float(tp1_val) if tp1_val else None,
                        take_profit_2=float(tp2_val) if tp2_val else None,
                        is_matched=False
                    )
                    db.session.add(log)
                    db.session.commit()
                    logger.info(f"[Dry-Run Mode] Saved LocalSignalLog to DB: {sig_type} at {current_price_live}")
                    
                    try:
                        append_signal_to_excel(
                            account_name=account_name,
                            sig_type=sig_type,
                            price=current_price_live,
                            quantity=float(size),
                            stop_loss=float(stop_loss_val) if stop_loss_val else None,
                            tp1=float(tp1_val) if tp1_val else None,
                            tp2=float(tp2_val) if tp2_val else None
                        )
                        logger.info(f"[Dry-Run Mode] Saved signal to Excel file: {sig_type}")
                    except Exception as xl_e:
                        logger.error(f"[Dry-Run Mode] Failed to save signal to Excel: {xl_e}")
                        
                except Exception as log_e:
                    logger.error(f"[Dry-Run Mode] Failed to save LocalSignalLog: {log_e}")
            
            # Return mock order success response
            return {
                "success": True,
                "result": {
                    "id": f"sim_{int(time.time() * 1000)}",
                    "average_fill_price": current_price_live,
                    "price": current_price_live
                }
            }
        client.place_order = mock_place_order
 
    # Get current ticker price for monitoring active trade exits
    ticker_data = client.get_ticker(symbol)
    if not ticker_data:
        logger.error(f"[{account_name}] Failed to fetch ticker for {symbol}.")
        return
    current_price = float(ticker_data.get("mark_price") or ticker_data.get("last_price") or last_close)
    
    # Verify position on exchange matches our DB state
    pos = client.get_position(product_id)
    exchange_size = 0.0
    if pos:
        try:
            exchange_size = float(pos.get("size", 0.0))
        except (ValueError, TypeError):
            exchange_size = 0.0
            
    # If exchange size is 0 but our state says we are in a position, we got stopped/liquidated/manually closed.
    if exchange_size == 0.0 and position_size != 0.0:
        logger.info(f"[{account_name}] Position on exchange is flat but local state is active. Resetting state to flat (manual exit detected).")
        manual_exit_dir = 1 if position_size > 0.0 else -1
        position_size = 0.0
        entry_price = None
        sl_dist = None
        tp1_price = None
        tp2_price = None
        tp1_hit = False
        tp2_hit = False
        current_sl = None
        manual_exit_detected = manual_exit_dir
        update_strategy_state(app, state_id, {
            "position_size": 0.0,
            "entry_price": None,
            "sl_dist": None,
            "tp1_price": None,
            "tp2_price": None,
            "tp1_hit": False,
            "tp2_hit": False,
            "current_sl": None,
            "manual_exit_detected": manual_exit_dir
        })
        
    # Check circuit breaker before any entry actions
    if is_circuit_broken:
        logger.warning(f"[{account_name}] Circuit breaker is broken. Enforcing flat state and skipping strategy runner entries.")
        if position_size != 0.0:
            # Emergency exit just in case
            client.place_order(product_id, size=abs(int(exchange_size)), side="sell" if exchange_size > 0 else "buy", order_type="market_order", reduce_only=True)
            update_strategy_state(app, state_id, {"position_size": 0.0})
            position_size = 0.0
    # 4. MONITOR ACTIVE POSITION
    if str(strategy_type).lower() == "ai_zl_fusion":
        if position_size > 0:  # Active Long Position
            if signals.get("exit_long") is not None and signals["exit_long"][candle_idx]:
                close_qty = abs(int(exchange_size))
                if close_qty > 0:
                    res = client.place_order(product_id, size=close_qty, side="sell", order_type="market_order", reduce_only=True)
                    if res.get("success"):
                        pnl_val = (current_price - (entry_price or current_price)) * close_qty * contract_val
                        send_strategy_notification(app, f"🟡 Local Strategy: Long Fusion Exit [{account_name}]", f"AI trend broke or market became choppy. Closed remaining {close_qty} contracts at {current_price:.2f} (PnL: {pnl_val:+.2f} USD).", symbol=symbol)
                update_strategy_state(app, state_id, {
                    "position_size": 0.0,
                    "entry_price": None,
                    "current_sl": None
                })
                position_size = 0.0
                return
        elif position_size < 0:  # Active Short Position
            if signals.get("exit_short") is not None and signals["exit_short"][candle_idx]:
                close_qty = abs(int(exchange_size))
                if close_qty > 0:
                    res = client.place_order(product_id, size=close_qty, side="buy", order_type="market_order", reduce_only=True)
                    if res.get("success"):
                        pnl_val = ((entry_price or current_price) - current_price) * close_qty * contract_val
                        send_strategy_notification(app, f"🟡 Local Strategy: Short Fusion Exit [{account_name}]", f"AI trend broke or market became choppy. Closed remaining {close_qty} contracts at {current_price:.2f} (PnL: {pnl_val:+.2f} USD).", symbol=symbol)
                update_strategy_state(app, state_id, {
                    "position_size": 0.0,
                    "entry_price": None,
                    "current_sl": None
                })
                position_size = 0.0
                return
    elif str(strategy_type).lower() != "zero_lag":
        if position_size > 0:  # Active Long Position
            entry_price_val = entry_price if entry_price else current_price
            # TP1 check (50% position close)
            if current_price >= tp1_price and not tp1_hit:
                close_qty = int(math.floor(position_size * 0.5))
                if close_qty > 0:
                    res = client.place_order(product_id, size=close_qty, side="sell", order_type="market_order", reduce_only=True)
                    if res.get("success"):
                        tp1_hit = True
                        updates = {"tp1_hit": True}
                        if use_be:
                            current_sl = max(current_sl, entry_price)
                            updates["current_sl"] = current_sl
                        update_strategy_state(app, state_id, updates)
                        pnl_val = (current_price - entry_price_val) * close_qty * contract_val
                        send_strategy_notification(app, f"🟢 Local Strategy: Long TP1 Hit [{account_name}]", f"Closed {close_qty} contracts at {current_price:.2f} (PnL: {pnl_val:+.2f} USD). SL moved to Break-Even ({current_sl:.2f}).", symbol=symbol)
                        
            # TP2 check (30% position close)
            if current_price >= tp2_price and not tp2_hit:
                close_qty = int(math.floor(position_size * 0.3))
                if close_qty > 0:
                    res = client.place_order(product_id, size=close_qty, side="sell", order_type="market_order", reduce_only=True)
                    if res.get("success"):
                        tp2_hit = True
                        update_strategy_state(app, state_id, {"tp2_hit": True})
                        pnl_val = (current_price - entry_price_val) * close_qty * contract_val
                        send_strategy_notification(app, f"🟢 Local Strategy: Long TP2 Hit [{account_name}]", f"Closed {close_qty} contracts at {current_price:.2f} (PnL: {pnl_val:+.2f} USD).", symbol=symbol)
    
            # Stop Loss check
            if current_price <= current_sl:
                close_qty = abs(int(exchange_size))
                if close_qty > 0:
                    res = client.place_order(product_id, size=close_qty, side="sell", order_type="market_order", reduce_only=True)
                    if res.get("success"):
                        pnl_val = (current_price - entry_price_val) * close_qty * contract_val
                        send_strategy_notification(app, f"🔴 Local Strategy: Long SL Hit [{account_name}]", f"Position stopped out. Closed remaining {close_qty} contracts at {current_price:.2f} (PnL: {pnl_val:+.2f} USD).", 15680580, symbol=symbol)
                update_strategy_state(app, state_id, {
                    "position_size": 0.0,
                    "entry_price": None,
                    "current_sl": None
                })
                return
                
            # ZLSMA Exit check
            if use_zlsma and last_close < signals["zlsma"][candle_idx]:
                close_qty = abs(int(exchange_size))
                if close_qty > 0:
                    res = client.place_order(product_id, size=close_qty, side="sell", order_type="market_order", reduce_only=True)
                    if res.get("success"):
                        pnl_val = (current_price - entry_price_val) * close_qty * contract_val
                        send_strategy_notification(app, f"🟡 Local Strategy: Long ZLSMA Flip Exit [{account_name}]", f"ZLSMA flipped bearish. Closed remaining {close_qty} contracts at {current_price:.2f} (PnL: {pnl_val:+.2f} USD).", 15549011, symbol=symbol)
                update_strategy_state(app, state_id, {"position_size": 0.0})
                return
                
            # Liquidity BSL Exit check
            if use_liq and tp1_hit and signals["bsl_created"][candle_idx]:
                close_qty = abs(int(exchange_size))
                if close_qty > 0:
                    res = client.place_order(product_id, size=close_qty, side="sell", order_type="market_order", reduce_only=True)
                    if res.get("success"):
                        pnl_val = (current_price - entry_price_val) * close_qty * contract_val
                        send_strategy_notification(app, f"🟡 Local Strategy: Long BSL Liquidity Exit [{account_name}]", f"BSL liquidity level created after TP1. Closed remaining {close_qty} contracts at {current_price:.2f} (PnL: {pnl_val:+.2f} USD).", 15549011, symbol=symbol)
                update_strategy_state(app, state_id, {"position_size": 0.0})
                return
     
        elif position_size < 0:  # Active Short Position
            entry_price_val = entry_price if entry_price else current_price
            # TP1 check (50% position close)
            if current_price <= tp1_price and not tp1_hit:
                close_qty = int(math.floor(abs(position_size) * 0.5))
                if close_qty > 0:
                    res = client.place_order(product_id, size=close_qty, side="buy", order_type="market_order", reduce_only=True)
                    if res.get("success"):
                        tp1_hit = True
                        updates = {"tp1_hit": True}
                        if use_be:
                            current_sl = min(current_sl, entry_price)
                            updates["current_sl"] = current_sl
                        update_strategy_state(app, state_id, updates)
                        pnl_val = (entry_price_val - current_price) * close_qty * contract_val
                        send_strategy_notification(app, f"🟢 Local Strategy: Short TP1 Hit [{account_name}]", f"Closed {close_qty} contracts at {current_price:.2f} (PnL: {pnl_val:+.2f} USD). SL moved to Break-Even ({current_sl:.2f}).", symbol=symbol)
                        
            # TP2 check (30% position close)
            if current_price <= tp2_price and not tp2_hit:
                close_qty = int(math.floor(abs(position_size) * 0.3))
                if close_qty > 0:
                    res = client.place_order(product_id, size=close_qty, side="buy", order_type="market_order", reduce_only=True)
                    if res.get("success"):
                        tp2_hit = True
                        update_strategy_state(app, state_id, {"tp2_hit": True})
                        pnl_val = (entry_price_val - current_price) * close_qty * contract_val
                        send_strategy_notification(app, f"🟢 Local Strategy: Short TP2 Hit [{account_name}]", f"Closed {close_qty} contracts at {current_price:.2f} (PnL: {pnl_val:+.2f} USD).", symbol=symbol)
    
            # Stop Loss check
            if current_price >= current_sl:
                close_qty = abs(int(exchange_size))
                if close_qty > 0:
                    res = client.place_order(product_id, size=close_qty, side="buy", order_type="market_order", reduce_only=True)
                    if res.get("success"):
                        pnl_val = (entry_price_val - current_price) * close_qty * contract_val
                        send_strategy_notification(app, f"🔴 Local Strategy: Short SL Hit [{account_name}]", f"Position stopped out. Closed remaining {close_qty} contracts at {current_price:.2f} (PnL: {pnl_val:+.2f} USD).", 15680580, symbol=symbol)
                update_strategy_state(app, state_id, {
                    "position_size": 0.0,
                    "entry_price": None,
                    "current_sl": None
                })
                return
                
            # ZLSMA Exit check
            if use_zlsma and last_close > signals["zlsma"][candle_idx]:
                close_qty = abs(int(exchange_size))
                if close_qty > 0:
                    res = client.place_order(product_id, size=close_qty, side="buy", order_type="market_order", reduce_only=True)
                    if res.get("success"):
                        pnl_val = (entry_price_val - current_price) * close_qty * contract_val
                        send_strategy_notification(app, f"🟡 Local Strategy: Short ZLSMA Flip Exit [{account_name}]", f"ZLSMA flipped bullish. Closed remaining {close_qty} contracts at {current_price:.2f} (PnL: {pnl_val:+.2f} USD).", 15549011, symbol=symbol)
                update_strategy_state(app, state_id, {"position_size": 0.0})
                return
                
            # Liquidity SSL Exit check
            if use_liq and tp1_hit and signals["ssl_created"][candle_idx]:
                close_qty = abs(int(exchange_size))
                if close_qty > 0:
                    res = client.place_order(product_id, size=close_qty, side="buy", order_type="market_order", reduce_only=True)
                    if res.get("success"):
                        pnl_val = (entry_price_val - current_price) * close_qty * contract_val
                        send_strategy_notification(app, f"🟡 Local Strategy: Short SSL Liquidity Exit [{account_name}]", f"SSL liquidity level created after TP1. Closed remaining {close_qty} contracts at {current_price:.2f} (PnL: {pnl_val:+.2f} USD).", 15549011, symbol=symbol)
                update_strategy_state(app, state_id, {"position_size": 0.0})
                return
 
    # 5. CHECK NEW ENTRIES (ONLY ON COMPLETED BAR TRANSITIONS)
    if last_signal_time == last_completed_time:
        return
        
    long_cond = signals["long_condition"][candle_idx]
    short_cond = signals["short_condition"][candle_idx]
    long_reentry = signals.get("long_reentry", np.zeros(len(df), dtype=bool))[candle_idx]
    short_reentry = signals.get("short_reentry", np.zeros(len(df), dtype=bool))[candle_idx]
    
    ai_trend_val = signals.get("ai_trend", np.ones(len(df), dtype=int))[candle_idx]
    zl_trend_val = signals.get("zl_trend", np.zeros(len(df), dtype=int))[candle_idx]
    is_choppy_val = signals.get("is_choppy", np.zeros(len(df), dtype=bool))[candle_idx]
    
    # Reset manual exit flag if trend conditions are no longer met
    if manual_exit_detected == 1:
        if ai_trend_val != 1 or zl_trend_val != 1 or is_choppy_val:
            logger.info(f"[{account_name}] Long trend broken/choppy after manual exit. Resetting manual exit state.")
            manual_exit_detected = 0
            update_strategy_state(app, state_id, {"manual_exit_detected": 0})
    elif manual_exit_detected == -1:
        if ai_trend_val != -1 or zl_trend_val != -1 or is_choppy_val:
            logger.info(f"[{account_name}] Short trend broken/choppy after manual exit. Resetting manual exit state.")
            manual_exit_detected = 0
            update_strategy_state(app, state_id, {"manual_exit_detected": 0})
            
    is_long_reentry_signal = (manual_exit_detected == 1 and long_reentry)
    is_short_reentry_signal = (manual_exit_detected == -1 and short_reentry)
    
    is_long_reversal = (long_cond and position_size < 0)
    is_short_reversal = (short_cond and position_size > 0)
    
    if is_long_reversal or is_short_reversal:
        logger.info(f"[{account_name}] Reversal signal detected on local strategy for {symbol}. Closing opposing position...")
        close_qty = abs(int(exchange_size))
        if close_qty > 0:
            client.place_order(product_id, size=close_qty, side="sell" if is_short_reversal else "buy", order_type="market_order", reduce_only=True)
            time.sleep(1.5)  # Let margin release
        position_size = 0.0
        update_strategy_state(app, state_id, {"position_size": 0.0})
        
    if (long_cond or is_long_reentry_signal) and position_size == 0.0:
        # Long Stop and SL Distance
        stop_px = signals["long_stop"][candle_idx]
        if np.isnan(stop_px):
            logger.warning(f"[{account_name}] Long stop is NaN for {symbol}. Skipping entry.")
            return
            
        sl_dist = max(last_close - stop_px, tick_size)
        
        # Calculate sizing based on account settings
        balance, asset = client.get_available_balance()
        lot_value_usd = last_close * contract_val
        
        if sizing_type == "fixed":
            buying_power = fixed_amount * account_leverage
            sizing_desc = f"Fixed Margin = {fixed_amount} {asset}"
        else:
            buying_power = balance * account_leverage * (balance_buffer_pct / 100.0)
            sizing_desc = f"Buffer = {balance_buffer_pct}%"
            
        qty_lots = int(math.floor(buying_power / lot_value_usd))
        
        # Enforce maximum buying power based on leverage
        max_buying_power = balance * account_leverage * 0.90
        max_qty_lots = int(math.floor(max_buying_power / lot_value_usd))
        qty_lots = min(qty_lots, max_qty_lots)
        
        if qty_lots <= 0:
            logger.warning(f"[{account_name}] Calculated long size is 0 lots for {symbol} (Balance: {balance:.2f}, Buying Power: {buying_power:.2f}).")
            update_strategy_state(app, state_id, {"last_signal_time": last_completed_time})
            return
            
        # Place Buy Order
        res = client.place_order(product_id, size=qty_lots, side="buy", order_type="market_order", reduce_only=False)
        if res.get("success"):
            update_strategy_state(app, state_id, {
                "position_size": float(qty_lots),
                "entry_price": last_close,
                "sl_dist": sl_dist,
                "tp1_price": last_close + sl_dist * tp1_rr,
                "tp2_price": last_close + sl_dist * tp2_rr,
                "tp1_hit": False,
                "tp2_hit": False,
                "current_sl": stop_px,
                "last_signal_time": last_completed_time,
                "manual_exit_detected": 0
            })
            
            reentry_tag = " (Re-Entry)" if is_long_reentry_signal else ""
            send_strategy_notification(
                app, 
                f"🟢 Local Strategy: Long Entry{reentry_tag} [{account_name}]", 
                f"Entered Long <b>{qty_lots} lots</b> on <b>{symbol}</b> at <b>{last_close:.2f}</b>\n"
                f"Stop Loss: {stop_px:.2f} (Dist: {sl_dist:.2f})\n"
                f"TP1: {last_close + sl_dist * tp1_rr:.2f} | TP2: {last_close + sl_dist * tp2_rr:.2f}",
                symbol=symbol
            )
            
            # Send Email Alert
            send_entry_email_alert(
                app,
                account_name=account_name,
                symbol=symbol,
                side="buy",
                qty_lots=qty_lots,
                price=last_close,
                stop_px=stop_px,
                sl_dist=sl_dist,
                is_reentry=is_long_reentry_signal
            )
            
            log_trade(app, symbol, "buy", "local_strategy", "success", f"Local Strategy Long Entry{reentry_tag} for {symbol}: {qty_lots} lots @ {last_close:.2f}")
        else:
            logger.error(f"[{account_name}] Long Entry order placement failed for {symbol}: {res}")
            
    elif (short_cond or is_short_reentry_signal) and position_size == 0.0:
        # Short Stop and SL Distance
        stop_px = signals["short_stop"][candle_idx]
        if np.isnan(stop_px):
            logger.warning(f"[{account_name}] Short stop is NaN for {symbol}. Skipping entry.")
            return
            
        sl_dist = max(stop_px - last_close, tick_size)
        
        # Calculate sizing based on account settings
        balance, asset = client.get_available_balance()
        lot_value_usd = last_close * contract_val
        
        if sizing_type == "fixed":
            buying_power = fixed_amount * account_leverage
            sizing_desc = f"Fixed Margin = {fixed_amount} {asset}"
        else:
            buying_power = balance * account_leverage * (balance_buffer_pct / 100.0)
            sizing_desc = f"Buffer = {balance_buffer_pct}%"
            
        qty_lots = int(math.floor(buying_power / lot_value_usd))
        
        # Enforce maximum buying power based on leverage
        max_buying_power = balance * account_leverage * 0.90
        max_qty_lots = int(math.floor(max_buying_power / lot_value_usd))
        qty_lots = min(qty_lots, max_qty_lots)
        
        if qty_lots <= 0:
            logger.warning(f"[{account_name}] Calculated short size is 0 lots for {symbol} (Balance: {balance:.2f}, Buying Power: {buying_power:.2f}).")
            update_strategy_state(app, state_id, {"last_signal_time": last_completed_time})
            return
            
        # Place Sell Order
        res = client.place_order(product_id, size=qty_lots, side="sell", order_type="market_order", reduce_only=False)
        if res.get("success"):
            update_strategy_state(app, state_id, {
                "position_size": -float(qty_lots),
                "entry_price": last_close,
                "sl_dist": sl_dist,
                "tp1_price": last_close - sl_dist * tp1_rr,
                "tp2_price": last_close - sl_dist * tp2_rr,
                "tp1_hit": False,
                "tp2_hit": False,
                "current_sl": stop_px,
                "last_signal_time": last_completed_time,
                "manual_exit_detected": 0
            })
            
            reentry_tag = " (Re-Entry)" if is_short_reentry_signal else ""
            send_strategy_notification(
                app, 
                f"🟢 Local Strategy: Short Entry{reentry_tag} [{account_name}]", 
                f"Entered Short <b>{qty_lots} lots</b> on <b>{symbol}</b> at <b>{last_close:.2f}</b>\n"
                f"Stop Loss: {stop_px:.2f} (Dist: {sl_dist:.2f})\n"
                f"TP1: {last_close - sl_dist * tp1_rr:.2f} | TP2: {last_close - sl_dist * tp2_rr:.2f}",
                symbol=symbol
            )
            
            # Send Email Alert
            send_entry_email_alert(
                app,
                account_name=account_name,
                symbol=symbol,
                side="sell",
                qty_lots=qty_lots,
                price=last_close,
                stop_px=stop_px,
                sl_dist=sl_dist,
                is_reentry=is_short_reentry_signal
            )
            
            log_trade(app, symbol, "sell", "local_strategy", "success", f"Local Strategy Short Entry{reentry_tag} for {symbol}: {qty_lots} lots @ {last_close:.2f}")
        else:
            logger.error(f"[{account_name}] Short Entry order placement failed for {symbol}: {res}")
            
    else:
        # No signal, just record that we evaluated this candle
        update_strategy_state(app, state_id, {"last_signal_time": last_completed_time})

def strategy_runner_loop(app):
    """Main strategy daemon loop that executes every 10 seconds for active accounts and symbols."""
    logger.info("Local Python Strategy Runner Thread started.")
    
    cycle_count = 0
    while True:
        try:
            # Periodically verify API connectivity (every 60 seconds)
            if cycle_count % 6 == 0:
                try:
                    from app import verify_api_connectivity
                    verify_api_connectivity()
                except Exception as conn_e:
                    logger.error(f"Failed to check API connectivity: {conn_e}")
            cycle_count += 1

            # Fetch settings inside database context
            strategy_type = "chandelier_exit"
            ce_length = 22
            ce_mult = 3.0
            zlsma_length = 32
            vol_length = 20
            vol_mult = 1.15
            zl_length = 70
            zl_mult = 1.2
            ai_speed = 14
            ai_atr_len = 14
            ai_atr_mult = 2.0
            resolution = "5"
            tp1_rr = 1.5
            tp2_rr = 3.0
            use_be = True
            use_liq = True
            use_zlsma = True
            symbols_list = [SYMBOL]
            
            with app.app_context():
                from models import GlobalSetting
                keys = [
                    "local_bot_symbols", "local_strategy_ce_length", "local_strategy_ce_mult",
                    "local_strategy_zlsma_length", "local_strategy_vol_length", "local_strategy_vol_mult",
                    "local_strategy_resolution", "local_strategy_tp1_rr", "local_strategy_tp2_rr",
                    "local_strategy_use_be", "local_strategy_use_liq_exit", "local_strategy_use_zlsma_exit",
                    "local_strategy_type", "local_strategy_zl_length", "local_strategy_zl_mult",
                    "local_strategy_ai_speed", "local_strategy_ai_atr_len", "local_strategy_ai_atr_mult"
                ]
                settings = GlobalSetting.query.filter(GlobalSetting.key.in_(keys)).all()
                s_dict = {s.key: s.value for s in settings}
                
                if "local_bot_symbols" in s_dict:
                    symbols_list = [s.strip() for s in s_dict["local_bot_symbols"].split(",") if s.strip()]
                
                try:
                    if "local_strategy_type" in s_dict: strategy_type = s_dict["local_strategy_type"]
                    if "local_strategy_ce_length" in s_dict: ce_length = int(s_dict["local_strategy_ce_length"])
                    if "local_strategy_ce_mult" in s_dict: ce_mult = float(s_dict["local_strategy_ce_mult"])
                    if "local_strategy_zlsma_length" in s_dict: zlsma_length = int(s_dict["local_strategy_zlsma_length"])
                    if "local_strategy_vol_length" in s_dict: vol_length = int(s_dict["local_strategy_vol_length"])
                    if "local_strategy_vol_mult" in s_dict: vol_mult = float(s_dict["local_strategy_vol_mult"])
                    if "local_strategy_zl_length" in s_dict: zl_length = int(s_dict["local_strategy_zl_length"])
                    if "local_strategy_zl_mult" in s_dict: zl_mult = float(s_dict["local_strategy_zl_mult"])
                    if "local_strategy_ai_speed" in s_dict: ai_speed = int(s_dict["local_strategy_ai_speed"])
                    if "local_strategy_ai_atr_len" in s_dict: ai_atr_len = int(s_dict["local_strategy_ai_atr_len"])
                    if "local_strategy_ai_atr_mult" in s_dict: ai_atr_mult = float(s_dict["local_strategy_ai_atr_mult"])
                    if "local_strategy_resolution" in s_dict: resolution = s_dict["local_strategy_resolution"]
                    if "local_strategy_tp1_rr" in s_dict: tp1_rr = float(s_dict["local_strategy_tp1_rr"])
                    if "local_strategy_tp2_rr" in s_dict: tp2_rr = float(s_dict["local_strategy_tp2_rr"])
                    if "local_strategy_use_be" in s_dict: use_be = s_dict["local_strategy_use_be"].lower() == "true"
                    if "local_strategy_use_liq_exit" in s_dict: use_liq = s_dict["local_strategy_use_liq_exit"].lower() == "true"
                    if "local_strategy_use_zlsma_exit" in s_dict: use_zlsma = s_dict["local_strategy_use_zlsma_exit"].lower() == "true"
                except Exception as e:
                    logger.error(f"Error parsing custom strategy settings: {e}")

            accounts_data = []
            with app.app_context():
                # Fetch active accounts with strategy enabled
                active_accounts = Account.query.filter_by(is_active=True, local_strategy_enabled=True).all()
                for account in active_accounts:
                    accounts_data.append({
                        "id": account.id,
                        "name": account.name,
                        "api_key": account.api_key,
                        "api_secret": account.api_secret,
                        "leverage": account.leverage,
                        "is_circuit_broken": account.is_circuit_broken,
                        "balance_buffer_pct": account.balance_buffer_pct,
                        "sizing_type": account.sizing_type,
                        "fixed_amount": account.fixed_amount
                    })
            
            for account_data in accounts_data:
                try:
                    client = DeltaClient(
                        api_key=account_data["api_key"],
                        api_secret=account_data["api_secret"],
                        base_url=app.config.get("BASE_URL", "https://api.delta.exchange")
                    )
                    
                    for symbol in symbols_list:
                        try:
                            run_strategy_for_account(
                                app, account_data, client, symbol=symbol,
                                strategy_type=strategy_type,
                                ce_length=ce_length, ce_mult=ce_mult, zlsma_length=zlsma_length,
                                vol_length=vol_length, vol_mult=vol_mult, tp1_rr=tp1_rr, tp2_rr=tp2_rr,
                                use_be=use_be, use_liq=use_liq, use_zlsma=use_zlsma, resolution=resolution,
                                zl_length=zl_length, zl_mult=zl_mult,
                                ai_speed=ai_speed, ai_atr_len=ai_atr_len, ai_atr_mult=ai_atr_mult
                            )
                        except Exception as sym_e:
                            logger.exception(f"Exception running strategy for {symbol} on account '{account_data['name']}': {sym_e}")
                            
                    # Trigger consecutive loss checks for the account
                    try:
                        from loss_analyzer import check_and_analyze_consecutive_losses
                        with app.app_context():
                            acc_model = Account.query.get(account_data["id"])
                            if acc_model:
                                check_and_analyze_consecutive_losses(app, acc_model)
                    except Exception as la_e:
                        logger.error(f"Error checking consecutive losses: {la_e}")
                except Exception as acc_e:
                    logger.exception(f"Exception running strategy on account '{account_data['name']}': {acc_e}")
                        
        except Exception as loop_e:
            logger.exception(f"Exception in main strategy runner loop: {loop_e}")
            
        time.sleep(10)
